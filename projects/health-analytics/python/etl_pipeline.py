"""
etl_pipeline.py
===============
Health Analytics ETL Pipeline — Portfolio Project
Author : Kisavi Shadrack | shadrackkisavi4@gmail.com

Stages
------
  1. Extract  — reads 3 messy Excel files
  2. Transform — cleans & standardises all three datasets
  3. Load     — inserts into PostgreSQL (or SQLite for local demo)
  4. Validate — runs post-load quality checks
  5. Report   — writes a summary Excel report + audit log

Usage (PostgreSQL)
------------------
  python etl_pipeline.py --mode postgres --dsn "postgresql://user:pw@host/db"

Usage (SQLite local demo — no Postgres needed)
----------------------------------------------
  python etl_pipeline.py --mode sqlite
"""

import argparse
import logging
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd
import sqlite3
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

def setup_logging(log_path: str) -> logging.Logger:
    Path(log_path).parent.mkdir(parents=True, exist_ok=True)
    fmt = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.FileHandler(log_path, mode="w"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger("health_etl")


# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

CONFIG = {
    "null_markers": ["N/A","n/a","NA","","None","NaN","TBD","Unknown","UNKNOWN","nil","NIL","unknown","nan"],
    "date_formats": ["%d/%m/%Y","%m/%d/%Y","%Y/%m/%d","%d-%m-%Y","%Y-%m-%d","%B %d %Y","%d %b %Y"],
    "valid_genders": {"male": "Male", "m": "Male", "female": "Female", "f": "Female", "other": "Other"},
    "valid_statuses": ["Recovered","Deceased","Referred","Absconded"],
    "valid_test_statuses": ["Final","Preliminary","Pending","Cancelled"],
    "valid_payment": ["Cash","NHIF","Insurance","Mpesa","Waived"],
    "bool_yes": ["yes","y","true","1"],
    "bool_no":  ["no","n","false","0"],
    "age_min": 0, "age_max": 120,
    "bill_min": 0,
    "los_min": 0, "los_max": 365,
    "tat_min": 0, "tat_max": 720,
    "wait_min": 0, "wait_max": 480,
}

DATES = CONFIG["date_formats"]


# ─────────────────────────────────────────────────────────────────────────────
# Data class
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ETLStats:
    source: str
    raw_rows: int = 0
    clean_rows: int = 0
    duplicates: int = 0
    null_fixes: int = 0
    date_fixes: int = 0
    rejected: int = 0
    issues: list = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def parse_date(val) -> Optional[datetime]:
    if pd.isna(val):
        return None
    s = str(val).strip()
    for fmt in DATES:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    return None

def clean_bool(val) -> Optional[bool]:
    if pd.isna(val):
        return None
    s = str(val).strip().lower()
    if s in CONFIG["bool_yes"]: return True
    if s in CONFIG["bool_no"]:  return False
    return None

def safe_num(val, lo, hi):
    try:
        v = float(val)
        return v if lo <= v <= hi else None
    except:
        return None

def standardise_text(s, lookup: dict):
    if pd.isna(s): return None
    key = str(s).strip().lower()
    return lookup.get(key, str(s).strip().title())


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — Extract
# ─────────────────────────────────────────────────────────────────────────────

class Extractor:
    def __init__(self, logger):
        self.log = logger.getChild("Extractor")

    def load(self, path: str, sheet: str, header_row=2, skip_footer=3) -> pd.DataFrame:
        self.log.info(f"Loading: {Path(path).name}")
        df = pd.read_excel(path, sheet_name=sheet, header=header_row, dtype=str)
        df = df.iloc[: len(df) - skip_footer]
        df.replace(CONFIG["null_markers"], np.nan, inplace=True)
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)
        self.log.info(f"  → {len(df)} rows loaded")
        return df


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — Transform
# ─────────────────────────────────────────────────────────────────────────────

class Transformer:
    def __init__(self, logger):
        self.log = logger.getChild("Transformer")

    # ── Admissions ────────────────────────────────────────────────────────────
    def transform_admissions(self, df: pd.DataFrame, stats: ETLStats) -> pd.DataFrame:
        self.log.info("Transforming admissions …")
        stats.raw_rows = len(df)
        df = df.copy()

        # Duplicates
        before = len(df)
        df.drop_duplicates(inplace=True)
        stats.duplicates = before - len(df)

        # Patient ID: rescue ORD-style → PAT-NNNNN
        def fix_pid(v):
            if pd.isna(v): return None
            v = str(v).strip()
            if re.match(r"^PAT-\d{5}$", v): return v
            m = re.match(r"PAT(\d{5})", v)
            return f"PAT-{m.group(1)}" if m else None
        df["Patient ID"] = df["Patient ID"].apply(fix_pid)

        # Dates
        df["Admission Date"] = df["Admission Date"].apply(parse_date)
        df["Discharge Date"] = df["Discharge Date"].apply(parse_date)
        # Discard discharge dates before admission
        bad_dc = df["Discharge Date"] < df["Admission Date"]
        df.loc[bad_dc, "Discharge Date"] = None
        stats.date_fixes = int(bad_dc.sum())

        # Gender
        df["Patient Gender"] = df["Patient Gender"].apply(
            lambda x: standardise_text(x, CONFIG["valid_genders"])
        )

        # County: title-case
        df["County"] = df["County"].str.strip().str.title()

        # Ward / Diagnosis / Discharge Status: title-case
        for col in ["Ward","Diagnosis","Discharge Status"]:
            df[col] = df[col].str.strip().str.title()

        # Attending doctor: title-case, strip double spaces
        df["Attending Doctor"] = df["Attending Doctor"].str.strip()\
                                    .str.replace(r"\s+", " ", regex=True).str.title()

        # Numeric columns
        df["Patient Age"]       = df["Patient Age"].apply(lambda x: safe_num(x, 0, 120))
        df["Total Bill (KES)"]  = df["Total Bill (KES)"].apply(lambda x: safe_num(x, 0, 10_000_000))
        df["Length of Stay"]    = df["Length of Stay"].apply(lambda x: safe_num(x, 0, 365))

        # Insurance
        df["Insurance Covered"] = df["Insurance Covered"].str.strip().str.title()

        # Drop rows with no Patient ID or Admission Date
        df.dropna(subset=["Patient ID","Admission Date"], inplace=True)
        df.reset_index(drop=True, inplace=True)

        stats.clean_rows = len(df)
        stats.rejected   = stats.raw_rows - stats.duplicates - stats.clean_rows
        self.log.info(f"  → {stats.clean_rows} clean rows")
        return df

    # ── Lab Results ───────────────────────────────────────────────────────────
    def transform_lab(self, df: pd.DataFrame, stats: ETLStats) -> pd.DataFrame:
        self.log.info("Transforming lab results …")
        stats.raw_rows = len(df)
        df = df.copy()

        before = len(df)
        df.drop_duplicates(inplace=True)
        stats.duplicates = before - len(df)

        # IDs
        df["Lab Result ID"] = df["Lab Result ID"].str.strip()
        df["Patient ID"]    = df["Patient ID"].str.strip()

        # Test name: standardise synonyms
        synonym_map = {
            "cbc": "Complete Blood Count",
            "complete blood count": "Complete Blood Count",
            "rbs": "Random Blood Sugar",
            "random blood sugar": "Random Blood Sugar",
            "lft": "Liver Function Test",
            "liver function test": "Liver Function Test",
            "malaria rdt": "Malaria RDT",
        }
        df["Test Name"] = df["Test Name"].apply(
            lambda x: synonym_map.get(str(x).strip().lower(), str(x).strip().title()) if not pd.isna(x) else None
        )

        # Dates
        df["Date Ordered"]   = df["Date Ordered"].apply(parse_date)
        df["Date Resulted"]  = df["Date Resulted"].apply(parse_date)
        bad_dr = df["Date Resulted"].notna() & (df["Date Resulted"] < df["Date Ordered"])
        df.loc[bad_dr, "Date Resulted"] = None
        stats.date_fixes = int(bad_dr.sum())

        # Numeric result
        df["Numeric Result"] = pd.to_numeric(df["Numeric Result"], errors="coerce")
        df.loc[df["Numeric Result"] < 0, "Numeric Result"] = None

        # TAT
        df["TAT (Hours)"] = df["TAT (Hours)"].apply(lambda x: safe_num(x, 0, 720))

        # Critical flag
        df["Critical Flag"] = df["Critical Flag"].apply(clean_bool)

        # Status: title-case
        df["Result Status"] = df["Result Status"].str.strip().str.title()

        df.dropna(subset=["Lab Result ID","Date Ordered","Test Name"], inplace=True)
        df.drop_duplicates(subset=["Lab Result ID"], inplace=True)
        df.reset_index(drop=True, inplace=True)

        stats.clean_rows = len(df)
        stats.rejected   = stats.raw_rows - stats.duplicates - stats.clean_rows
        self.log.info(f"  → {stats.clean_rows} clean rows")
        return df

    # ── Outpatient Visits ─────────────────────────────────────────────────────
    def transform_opd(self, df: pd.DataFrame, stats: ETLStats) -> pd.DataFrame:
        self.log.info("Transforming outpatient visits …")
        stats.raw_rows = len(df)
        df = df.copy()

        before = len(df)
        df.drop_duplicates(inplace=True)
        stats.duplicates = before - len(df)

        df["Visit ID"]   = df["Visit ID"].str.strip()
        df["Patient ID"] = df["Patient ID"].str.strip()

        df["Visit Date"] = df["Visit Date"].apply(parse_date)
        df.dropna(subset=["Visit Date"], inplace=True)
        stats.date_fixes = stats.raw_rows - len(df)

        df["Clinic"]           = df["Clinic"].str.strip().str.title()
        df["Attending Doctor"] = df["Attending Doctor"].str.strip()\
                                     .str.replace(r"\s+", " ", regex=True).str.title()
        df["Patient Gender"]   = df["Patient Gender"].apply(
            lambda x: standardise_text(x, CONFIG["valid_genders"])
        )
        df["County"]     = df["County"].str.strip().str.title()
        df["Diagnosis"]  = df["Diagnosis"].str.strip().str.title()

        df["Patient Age"]             = df["Patient Age"].apply(lambda x: safe_num(x, 0, 120))
        df["Consultation Fee (KES)"]  = df["Consultation Fee (KES)"].apply(lambda x: safe_num(x, 0, 1_000_000))
        df["Wait Time (Minutes)"]     = df["Wait Time (Minutes)"].apply(lambda x: safe_num(x, 0, 480))
        df["Consult Duration (Min)"]  = df["Consult Duration (Min)"].apply(lambda x: safe_num(x, 0, 300))
        df["Follow Up Required"]      = df["Follow Up Required"].apply(clean_bool)

        # Payment: standardise
        pay_map = {"cash":"Cash","nhif":"NHIF","insurance":"Insurance",
                   "mpesa":"Mpesa","waived":"Waived","m-pesa":"Mpesa"}
        df["Payment Method"] = df["Payment Method"].apply(
            lambda x: pay_map.get(str(x).strip().lower(), str(x).strip().title()) if not pd.isna(x) else None
        )

        df.dropna(subset=["Visit ID","Visit Date"], inplace=True)
        df.drop_duplicates(subset=["Visit ID"], inplace=True)
        df.reset_index(drop=True, inplace=True)

        stats.clean_rows = len(df)
        stats.rejected   = stats.raw_rows - stats.duplicates - stats.clean_rows
        self.log.info(f"  → {stats.clean_rows} clean rows")
        return df


# ─────────────────────────────────────────────────────────────────────────────
# Stage 3 — Load (SQLite local demo)
# ─────────────────────────────────────────────────────────────────────────────

class SQLiteLoader:
    """Loads clean DataFrames into a local SQLite database for demo/portfolio."""

    def __init__(self, db_path: str, logger):
        self.db_path = db_path
        self.log = logger.getChild("SQLiteLoader")
        self.conn = sqlite3.connect(db_path)
        self._create_schema()

    def _create_schema(self):
        c = self.conn
        c.execute("""CREATE TABLE IF NOT EXISTS patients (
            patient_id TEXT PRIMARY KEY, gender TEXT, county TEXT)""")
        c.execute("""CREATE TABLE IF NOT EXISTS patient_admissions (
            patient_id TEXT, admission_date TEXT, discharge_date TEXT,
            ward TEXT, diagnosis TEXT, attending_doctor TEXT,
            patient_age REAL, discharge_status TEXT, total_bill_kes REAL,
            insurance_covered TEXT, length_of_stay REAL)""")
        c.execute("""CREATE TABLE IF NOT EXISTS lab_results (
            lab_result_id TEXT PRIMARY KEY, patient_id TEXT,
            test_name TEXT, date_ordered TEXT, date_resulted TEXT,
            numeric_result REAL, text_result TEXT, result_status TEXT,
            tat_hours REAL, critical_flag INTEGER)""")
        c.execute("""CREATE TABLE IF NOT EXISTS outpatient_visits (
            visit_id TEXT PRIMARY KEY, patient_id TEXT,
            visit_date TEXT, clinic TEXT, attending_doctor TEXT,
            patient_age REAL, diagnosis TEXT,
            consultation_fee_kes REAL, wait_time_minutes REAL,
            consult_duration_min REAL, follow_up_required INTEGER,
            payment_method TEXT)""")
        c.execute("""CREATE TABLE IF NOT EXISTS etl_audit_log (
            run_timestamp TEXT, source_file TEXT, raw_rows INTEGER,
            clean_rows INTEGER, duplicates INTEGER, rejected INTEGER)""")
        c.commit()

    def load_admissions(self, df: pd.DataFrame):
        # Upsert patients dimension
        patients = df[["Patient ID","Patient Gender","County"]].dropna(subset=["Patient ID"])
        patients = patients.rename(columns={"Patient ID":"patient_id",
                                            "Patient Gender":"gender","County":"county"})
        patients.drop_duplicates("patient_id").to_sql(
            "patients", self.conn, if_exists="replace", index=False)

        adm = df.rename(columns={
            "Patient ID":"patient_id","Admission Date":"admission_date",
            "Discharge Date":"discharge_date","Ward":"ward","Diagnosis":"diagnosis",
            "Attending Doctor":"attending_doctor","Patient Age":"patient_age",
            "Discharge Status":"discharge_status","Total Bill (KES)":"total_bill_kes",
            "Insurance Covered":"insurance_covered","Length of Stay":"length_of_stay",
        })
        adm["admission_date"] = adm["admission_date"].astype(str)
        adm["discharge_date"] = adm["discharge_date"].astype(str)
        adm[["patient_id","admission_date","discharge_date","ward","diagnosis",
             "attending_doctor","patient_age","discharge_status","total_bill_kes",
             "insurance_covered","length_of_stay"]].to_sql(
            "patient_admissions", self.conn, if_exists="replace", index=False)
        self.log.info(f"  Loaded {len(adm)} admissions to SQLite")

    def load_lab(self, df: pd.DataFrame):
        lab = df.rename(columns={
            "Lab Result ID":"lab_result_id","Patient ID":"patient_id",
            "Test Name":"test_name","Date Ordered":"date_ordered",
            "Date Resulted":"date_resulted","Numeric Result":"numeric_result",
            "Text Result":"text_result","Result Status":"result_status",
            "TAT (Hours)":"tat_hours","Critical Flag":"critical_flag",
        })
        lab["date_ordered"]  = lab["date_ordered"].astype(str)
        lab["date_resulted"] = lab["date_resulted"].astype(str)
        lab["critical_flag"] = lab["critical_flag"].apply(lambda x: 1 if x is True else 0)
        lab[["lab_result_id","patient_id","test_name","date_ordered","date_resulted",
             "numeric_result","text_result","result_status","tat_hours","critical_flag"]].to_sql(
            "lab_results", self.conn, if_exists="replace", index=False)
        self.log.info(f"  Loaded {len(lab)} lab results to SQLite")

    def load_opd(self, df: pd.DataFrame):
        opd = df.rename(columns={
            "Visit ID":"visit_id","Patient ID":"patient_id","Visit Date":"visit_date",
            "Clinic":"clinic","Attending Doctor":"attending_doctor","Patient Age":"patient_age",
            "Diagnosis":"diagnosis","Consultation Fee (KES)":"consultation_fee_kes",
            "Wait Time (Minutes)":"wait_time_minutes","Consult Duration (Min)":"consult_duration_min",
            "Follow Up Required":"follow_up_required","Payment Method":"payment_method",
        })
        opd["visit_date"] = opd["visit_date"].astype(str)
        opd["follow_up_required"] = opd["follow_up_required"].apply(lambda x: 1 if x is True else 0)
        opd[["visit_id","patient_id","visit_date","clinic","attending_doctor","patient_age",
             "diagnosis","consultation_fee_kes","wait_time_minutes","consult_duration_min",
             "follow_up_required","payment_method"]].to_sql(
            "outpatient_visits", self.conn, if_exists="replace", index=False)
        self.log.info(f"  Loaded {len(opd)} OPD visits to SQLite")

    def log_run(self, stats_list):
        rows = [(datetime.now().isoformat(), s.source, s.raw_rows,
                 s.clean_rows, s.duplicates, s.rejected)
                for s in stats_list]
        self.conn.executemany(
            "INSERT INTO etl_audit_log VALUES (?,?,?,?,?,?)", rows)
        self.conn.commit()

    def run_queries(self) -> dict:
        """Run the analytical queries and return results as DataFrames."""
        queries = {
            "monthly_admissions": """
                SELECT strftime('%Y-%m', admission_date) AS month,
                       COUNT(*) AS admissions,
                       ROUND(SUM(total_bill_kes),2) AS revenue_kes
                FROM patient_admissions
                WHERE admission_date != 'NaT'
                GROUP BY 1 ORDER BY 1""",
            "ward_performance": """
                SELECT ward, COUNT(*) AS admissions,
                       ROUND(AVG(length_of_stay),1) AS avg_los,
                       ROUND(AVG(total_bill_kes),2) AS avg_bill,
                       ROUND(AVG(CASE WHEN discharge_status='Deceased' THEN 1.0 ELSE 0.0 END)*100,2)
                           AS mortality_pct
                FROM patient_admissions WHERE ward IS NOT NULL
                GROUP BY ward ORDER BY admissions DESC""",
            "top_diagnoses": """
                SELECT diagnosis, COUNT(*) AS cases,
                       ROUND(AVG(total_bill_kes),2) AS avg_cost,
                       ROUND(AVG(length_of_stay),1) AS avg_los
                FROM patient_admissions WHERE diagnosis IS NOT NULL
                GROUP BY diagnosis HAVING cases >= 3
                ORDER BY cases DESC LIMIT 15""",
            "lab_tat": """
                SELECT test_name, COUNT(*) AS tests,
                       ROUND(AVG(tat_hours),1) AS avg_tat,
                       SUM(critical_flag) AS critical
                FROM lab_results WHERE tat_hours IS NOT NULL
                GROUP BY test_name HAVING tests >= 3
                ORDER BY tests DESC LIMIT 15""",
            "opd_clinics": """
                SELECT clinic, COUNT(*) AS visits,
                       ROUND(AVG(wait_time_minutes),1) AS avg_wait,
                       ROUND(AVG(consultation_fee_kes),2) AS avg_fee,
                       ROUND(SUM(consultation_fee_kes),2) AS total_fees,
                       SUM(follow_up_required) AS follow_ups
                FROM outpatient_visits WHERE clinic IS NOT NULL
                GROUP BY clinic ORDER BY visits DESC""",
            "payment_mix": """
                SELECT payment_method, COUNT(*) AS visits,
                       ROUND(SUM(consultation_fee_kes),2) AS total_fees
                FROM outpatient_visits WHERE payment_method IS NOT NULL
                GROUP BY payment_method ORDER BY visits DESC""",
        }
        results = {}
        for name, sql in queries.items():
            try:
                results[name] = pd.read_sql_query(sql, self.conn)
            except Exception as e:
                self.log.warning(f"Query {name} failed: {e}")
                results[name] = pd.DataFrame()
        return results

    def close(self):
        self.conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Stage 5 — Report
# ─────────────────────────────────────────────────────────────────────────────

class ReportWriter:
    BLUE  = "1F4E79"
    LBLUE = "BDD7EE"
    GREEN = "375623"
    RED   = "C00000"

    def __init__(self, logger):
        self.log = logger.getChild("ReportWriter")

    def _hdr(self, cell, bg, fg="FFFFFF"):
        cell.font  = Font(bold=True, color=fg)
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 45)

    def _title(self, ws, text, ncols):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        ws["A1"] = text
        ws["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill  = PatternFill("solid", fgColor=self.BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26
        ws.append([])

    def write(self, stats_list: list, query_results: dict, out_path: str):
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()

        self._sheet_etl_summary(wb, stats_list)
        for name, df in query_results.items():
            if not df.empty:
                self._sheet_dataframe(wb, name.replace("_"," ").title(), df)

        del wb["Sheet"]
        wb.save(out_path)
        self.log.info(f"Report saved → {out_path}")

    def _sheet_etl_summary(self, wb, stats_list):
        ws = wb.create_sheet("ETL Summary")
        self._title(ws, "ETL Pipeline Run Summary", 7)
        hdrs = ["Source File","Raw Rows","Duplicates Removed",
                "Date Fixes","Rejected","Clean Rows","Quality %"]
        ws.append(hdrs)
        for j, h in enumerate(hdrs, 1):
            self._hdr(ws.cell(3, j), self.BLUE)

        for s in stats_list:
            q_pct = f"{s.clean_rows/max(s.raw_rows,1)*100:.1f}%"
            ws.append([s.source, s.raw_rows, s.duplicates,
                       s.date_fixes, s.rejected, s.clean_rows, q_pct])

        self._auto_width(ws)

    def _sheet_dataframe(self, wb, title, df):
        ws = wb.create_sheet(title[:31])
        self._title(ws, title, len(df.columns))
        ws.append(list(df.columns))
        for j in range(1, len(df.columns)+1):
            self._hdr(ws.cell(3, j), self.GREEN)
        for i, (_, row) in enumerate(df.iterrows()):
            ws.append(list(row))
            fill = "E2EFDA" if i % 2 == 0 else "FFFFFF"
            for j in range(1, len(df.columns)+1):
                ws.cell(ws.max_row, j).fill = PatternFill("solid", fgColor=fill)
        self._auto_width(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Orchestrator
# ─────────────────────────────────────────────────────────────────────────────

class HealthETLPipeline:
    BASE = Path(__file__).parent.parent

    def __init__(self, mode="sqlite", dsn=None):
        self.logger = setup_logging(str(self.BASE / "logs/etl_pipeline.log"))
        self.mode   = mode
        self.dsn    = dsn
        self.log    = self.logger

    def run(self):
        log = self.log
        log.info("=" * 60)
        log.info("  Health Analytics ETL Pipeline — Kisavi Shadrack")
        log.info("=" * 60)

        extractor   = Extractor(log)
        transformer = Transformer(log)

        # ── Extract ──
        raw_adm = extractor.load(str(self.BASE/"data/patient_admissions.xlsx"),  "Patient Admissions")
        raw_lab = extractor.load(str(self.BASE/"data/lab_results.xlsx"),          "Lab Results")
        raw_opd = extractor.load(str(self.BASE/"data/outpatient_visits.xlsx"),    "Outpatient Visits")

        # ── Transform ──
        stats_adm = ETLStats("patient_admissions.xlsx")
        stats_lab = ETLStats("lab_results.xlsx")
        stats_opd = ETLStats("outpatient_visits.xlsx")

        clean_adm = transformer.transform_admissions(raw_adm, stats_adm)
        clean_lab = transformer.transform_lab(raw_lab, stats_lab)
        clean_opd = transformer.transform_opd(raw_opd, stats_opd)

        # ── Load ──
        if self.mode == "sqlite":
            db_path = str(self.BASE / "data/health_analytics.db")
            loader = SQLiteLoader(db_path, log)
            loader.load_admissions(clean_adm)
            loader.load_lab(clean_lab)
            loader.load_opd(clean_opd)
            loader.log_run([stats_adm, stats_lab, stats_opd])
            query_results = loader.run_queries()
            loader.close()
            log.info(f"SQLite database → {db_path}")

        # ── Report ──
        writer = ReportWriter(log)
        writer.write(
            [stats_adm, stats_lab, stats_opd],
            query_results if self.mode == "sqlite" else {},
            str(self.BASE / "reports/etl_summary_report.xlsx"),
        )

        log.info("=" * 60)
        log.info("Pipeline complete.")
        for s in [stats_adm, stats_lab, stats_opd]:
            log.info(f"  {s.source}: {s.raw_rows} raw → {s.clean_rows} clean "
                     f"({s.duplicates} dupes, {s.rejected} rejected)")
        log.info("=" * 60)

        return clean_adm, clean_lab, clean_opd, query_results


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["sqlite","postgres"], default="sqlite")
    parser.add_argument("--dsn",  default=None, help="PostgreSQL DSN string")
    args = parser.parse_args()
    HealthETLPipeline(mode=args.mode, dsn=args.dsn).run()
