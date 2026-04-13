"""
generate_health_data.py
=======================
Generates three messy, realistic healthcare Excel datasets:
  1. patient_admissions.xlsx   — Hospital inpatient admissions
  2. lab_results.xlsx          — Laboratory test results
  3. outpatient_visits.xlsx    — Outpatient clinic visits

Simulates real-world data quality issues found in exported HMIS/LIMS systems:
duplicates, nulls, mixed date formats, inconsistent casing, impossible values,
referential mismatches, merged-cell artefacts, and trailing footer noise.
"""

import random
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from pathlib import Path

random.seed(7)
np.random.seed(7)

OUT = Path("/home/claude/health_analytics_pipeline/data")
OUT.mkdir(exist_ok=True)

# ── Lookup pools ──────────────────────────────────────────────────────────────

COUNTIES   = ["Nairobi","nairobi","NAIROBI","Mombasa","mombasa","Kisumu","Nakuru","Eldoret","ELDORET","Thika"]
GENDERS    = ["Male","male","MALE","Female","female","F","M","Other","Unknown",None,"N/A"]
WARDS      = ["Cardiology","cardiology","ICU","icu","Maternity","MATERNITY","Pediatrics",
               "pediatrics","General","Surgery","SURGERY","Orthopedics","Oncology"]
DIAGNOSES  = ["Malaria","malaria","MALARIA","Hypertension","hypertension","Diabetes Mellitus",
               "diabetes mellitus","Pneumonia","PNEUMONIA","Typhoid Fever","typhoid fever",
               "HIV/AIDS","Tuberculosis","tuberculosis","Anaemia","Heart Failure","Appendicitis",
               "Fracture","Diarrhoea","Asthma","Preeclampsia","Sepsis"]
DISCHARGE  = ["Recovered","recovered","RECOVERED","Referred","referred","Deceased","deceased",
               "AEOR","Absconded","absconded",None,"N/A",""]
CLINICS    = ["Cardiology Clinic","cardiology clinic","General OPD","general opd","GENERAL OPD",
               "Antenatal Clinic","antenatal clinic","HIV Clinic","Dental","Eye Clinic",
               "Nutrition Clinic","ENT"]
TEST_NAMES = ["Complete Blood Count","complete blood count","CBC","Malaria RDT","malaria rdt",
               "Random Blood Sugar","random blood sugar","RBS","Urinalysis","urinalysis",
               "Liver Function Test","LFT","Creatinine","creatinine","HIV Rapid Test",
               "Tuberculin Test","Widal Test","Urine Culture","Blood Culture","HbA1c","CD4 Count"]
TEST_STATUS= ["Final","final","FINAL","Preliminary","preliminary","Pending","pending","Cancelled",None,"N/A"]
DOCTORS    = ["Dr. Otieno","DR. OTIENO","dr. otieno","Dr. Wanjiku","Dr. Kamau","dr. kamau",
               "Dr. Akinyi","Dr. Mwangi","dr. mwangi","Dr. Hassan","Unknown",None]
NULL_VALS  = ["N/A","n/a","NA","","None","NaN","TBD","Unknown","UNKNOWN","nil","NIL"]

def rand_date(start, end):
    return start + timedelta(days=random.randint(0, (end - start).days))

def bad_date(d):
    fmts = ["%d/%m/%Y","%m/%d/%Y","%Y/%m/%d","%d-%m-%Y","%B %d %Y","%d %b %Y"]
    bads = ["32/01/2023","13/13/2023","2023-99-01","Jan 2023","TBD","","N/A",None,"00/00/0000"]
    return random.choice(bads) if random.random() < 0.12 else d.strftime(random.choice(fmts))

def patient_id(i):
    if random.random() < 0.05: return None
    if random.random() < 0.05: return f"PAT{i:05d}"   # missing dash
    return f"PAT-{i:05d}"

def visit_id(i, prefix="V"):
    if random.random() < 0.04: return None
    return f"{prefix}-{i:06d}"

START = datetime(2023, 1, 1)
END   = datetime(2024, 6, 30)

# ─────────────────────────────────────────────────────────────────────────────
# 1. PATIENT ADMISSIONS
# ─────────────────────────────────────────────────────────────────────────────
def make_admissions(n=400):
    rows = []
    for i in range(1, n + 1):
        adm_date = rand_date(START, END)
        los = random.randint(1, 30)
        dis_date = adm_date + timedelta(days=los)

        # Occasionally swap or corrupt discharge date
        if random.random() < 0.05:
            dis_date_str = bad_date(adm_date - timedelta(days=2))  # before admission!
        elif random.random() < 0.07:
            dis_date_str = bad_date(dis_date)
        else:
            dis_date_str = dis_date.strftime("%d/%m/%Y")

        age = random.randint(0, 95) if random.random() > 0.05 else random.choice([-3, 999, None, "N/A", "adult"])
        bill = round(random.uniform(500, 120000), 2) if random.random() > 0.06 else random.choice([None, -500, 0, "N/A", "waived"])

        rows.append({
            "Patient ID":        patient_id(i),
            "Admission Date":    bad_date(adm_date),
            "Discharge Date":    dis_date_str,
            "Ward":              random.choice(WARDS),
            "Diagnosis":         random.choice(DIAGNOSES),
            "Attending Doctor":  random.choice(DOCTORS),
            "Patient Gender":    random.choice(GENDERS),
            "Patient Age":       age,
            "County":            random.choice(COUNTIES),
            "Discharge Status":  random.choice(DISCHARGE),
            "Total Bill (KES)":  bill,
            "Insurance Covered": random.choice(["Yes","No","yes","no","YES","NO",None,"N/A","Partial","partial"]),
            "Length of Stay":    los if random.random() > 0.08 else random.choice([None, -1, 0, "N/A", "unknown"]),
        })

    df = pd.DataFrame(rows)
    dupes = df.sample(20, random_state=3)
    df = pd.concat([df, dupes]).sample(frac=1, random_state=7).reset_index(drop=True)
    return df

# ─────────────────────────────────────────────────────────────────────────────
# 2. LAB RESULTS
# ─────────────────────────────────────────────────────────────────────────────
def make_lab_results(n=500):
    rows = []
    for i in range(1, n + 1):
        ordered = rand_date(START, END)
        tat_hrs = random.randint(1, 72)
        resulted = ordered + timedelta(hours=tat_hrs)

        # numeric result — sometimes text, sometimes impossible
        if random.random() < 0.06:
            result_val = random.choice(["Positive","Negative","positive","negative","reactive",
                                        "non-reactive","pending","N/A",None,"error"])
            result_num = None
        else:
            result_num = round(random.uniform(0.1, 500), 2)
            if random.random() < 0.04:
                result_num = random.choice([-99, 99999, None, "N/A", ""])
            result_val = None

        rows.append({
            "Lab Result ID":     visit_id(i, "LR"),
            "Patient ID":        f"PAT-{random.randint(1,400):05d}",
            "Test Name":         random.choice(TEST_NAMES),
            "Date Ordered":      bad_date(ordered),
            "Date Resulted":     bad_date(resulted) if random.random() > 0.08 else random.choice([None,"N/A",""]),
            "Numeric Result":    result_num,
            "Text Result":       result_val,
            "Reference Range":   random.choice(["3.5-5.0","<7.0","Negative","0-200","70-100","N/A",None,""]),
            "Result Status":     random.choice(TEST_STATUS),
            "Lab Technician":    random.choice(DOCTORS),
            "TAT (Hours)":       tat_hrs if random.random() > 0.07 else random.choice([None, -5, 0, 999, "N/A"]),
            "Critical Flag":     random.choice(["Yes","No","yes","no",None,"N/A","Y","N"]),
            "Equipment ID":      random.choice([f"EQ-{x:03d}" for x in range(1,10)] + [None,"N/A","Unknown"]),
        })

    df = pd.DataFrame(rows)
    dupes = df.sample(25, random_state=5)
    df = pd.concat([df, dupes]).sample(frac=1, random_state=7).reset_index(drop=True)
    return df

# ─────────────────────────────────────────────────────────────────────────────
# 3. OUTPATIENT VISITS
# ─────────────────────────────────────────────────────────────────────────────
def make_outpatient(n=600):
    rows = []
    for i in range(1, n + 1):
        visit_date = rand_date(START, END)
        wait_min = random.randint(5, 180)
        consult_min = random.randint(5, 60)

        age = random.randint(0, 90) if random.random() > 0.05 else random.choice([-1, 150, None, "N/A", "child"])
        fee = round(random.uniform(200, 5000), 2) if random.random() > 0.06 else random.choice([None, -200, 0, "N/A", "free"])

        rows.append({
            "Visit ID":              visit_id(i, "OPD"),
            "Patient ID":            f"PAT-{random.randint(1,400):05d}",
            "Visit Date":            bad_date(visit_date),
            "Clinic":                random.choice(CLINICS),
            "Attending Doctor":      random.choice(DOCTORS),
            "Patient Gender":        random.choice(GENDERS),
            "Patient Age":           age,
            "County":                random.choice(COUNTIES),
            "Chief Complaint":       random.choice(DIAGNOSES + [None,"N/A",""]),
            "Diagnosis":             random.choice(DIAGNOSES + [None,"N/A",""]),
            "Consultation Fee (KES)":fee,
            "Wait Time (Minutes)":   wait_min if random.random() > 0.07 else random.choice([None,-10,0,"N/A"]),
            "Consult Duration (Min)":consult_min if random.random() > 0.07 else random.choice([None,-5,0,"N/A"]),
            "Follow Up Required":    random.choice(["Yes","No","yes","no","YES","NO",None,"N/A","Maybe"]),
            "Payment Method":        random.choice(["Cash","NHIF","Insurance","nhif","cash","CASH","Mpesa","mpesa","MPESA",None,"N/A","Waived"]),
        })

    df = pd.DataFrame(rows)
    dupes = df.sample(30, random_state=9)
    df = pd.concat([df, dupes]).sample(frac=1, random_state=7).reset_index(drop=True)
    return df

# ─────────────────────────────────────────────────────────────────────────────
# Save with Excel noise
# ─────────────────────────────────────────────────────────────────────────────
def save_messy_excel(df, filename, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Fake system-export title row
    ncols = len(df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = f"EXPORTED FROM HMIS — {sheet_title.upper()} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.append([])  # blank row 2

    # Headers on row 3
    ws.append(list(df.columns))
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")

    # Data
    for _, row in df.iterrows():
        ws.append(list(row))

    # Footer noise
    ws.append([])
    ws.append(["*** END OF REPORT ***", None, None, "Records:", len(df)])
    ws.append(["Generated by: MedSoft HMIS v3.2", None, None, None, "CONFIDENTIAL"])

    wb.save(OUT / filename)
    print(f"  Saved {filename}  ({len(df)} rows)")

if __name__ == "__main__":
    print("Generating messy health datasets …")
    save_messy_excel(make_admissions(),   "patient_admissions.xlsx",  "Patient Admissions")
    save_messy_excel(make_lab_results(),  "lab_results.xlsx",         "Lab Results")
    save_messy_excel(make_outpatient(),   "outpatient_visits.xlsx",   "Outpatient Visits")
    print("Done.")
