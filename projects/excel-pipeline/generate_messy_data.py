"""
generate_messy_data.py
Generates a realistic messy Excel dataset (Sales & Orders) for portfolio demo.
Simulates real-world data quality issues: duplicates, nulls, mixed types,
inconsistent casing, bad dates, outliers, merged-cell-like blanks, etc.
"""

import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font

random.seed(42)
np.random.seed(42)

REGIONS = ["Nairobi", "NAIROBI", "nairobi", "Mombasa", "mombasa", "Kisumu", "Kisumu", "Eldoret", "ELDORET"]
PRODUCTS = ["Laptop", "laptop", "LAPTOP", "Phone", "Smartphone", "phone", "Tablet", "tablet", "Headphones", "Monitor", "monitor"]
STATUSES = ["Completed", "completed", "COMPLETED", "Pending", "pending", "Cancelled", "cancelled", "N/A", None, ""]
SALES_REPS = ["Alice Mwangi", "ALICE MWANGI", "alice mwangi", "Brian Otieno", "Brian  Otieno",
              "Carol Wanjiku", "carol wanjiku", "David Kimani", "david kimani", None, "N/A", "Unknown"]

def random_date(start, end):
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))

def make_bad_date():
    choices = [
        "13/25/2023", "2023-31-12", "Jan 2023", "2023", "32-01-2024",
        "15-15-2023", "N/A", "TBD", "", None
    ]
    return random.choice(choices)

def generate_order_id(i):
    if random.random() < 0.05:
        return None
    if random.random() < 0.05:
        return f"ORD{i:04d}"  # missing prefix dash
    return f"ORD-{i:04d}"

n = 300
start_date = datetime(2023, 1, 1)
end_date = datetime(2024, 6, 30)

rows = []
for i in range(1, n + 1):
    order_id = generate_order_id(i)

    # Occasionally blank out region (mimics merged cells)
    region = random.choice(REGIONS) if random.random() > 0.08 else None

    product = random.choice(PRODUCTS)
    rep = random.choice(SALES_REPS)
    status = random.choice(STATUSES)

    # Quantity: some bad values
    if random.random() < 0.05:
        qty = random.choice([-5, 0, "N/A", None, "five", ""])
    else:
        qty = random.randint(1, 50)

    # Unit price: some bad values, some outliers
    if random.random() < 0.05:
        price = random.choice([None, "N/A", -999, 0, "free", ""])
    elif random.random() < 0.02:
        price = random.uniform(50000, 200000)  # outlier
    else:
        price = round(random.uniform(500, 80000), 2)

    # Revenue: sometimes inconsistent with qty * price, sometimes missing
    if random.random() < 0.1:
        revenue = None
    elif random.random() < 0.05:
        revenue = round(random.uniform(100, 999999), 2)  # wrong value
    else:
        try:
            revenue = round(float(qty) * float(price), 2)
        except:
            revenue = None

    # Date: mix of good and bad
    if random.random() < 0.12:
        order_date = make_bad_date()
    else:
        order_date = random_date(start_date, end_date).strftime("%d/%m/%Y")

    # Customer email: some malformed
    if random.random() < 0.08:
        email = random.choice(["not-an-email", "missing@", "@nodomain.com", None, "N/A", ""])
    else:
        names = ["john", "mary", "james", "wanjiru", "otieno", "fatuma"]
        domains = ["gmail.com", "yahoo.com", "outlook.com", "company.co.ke"]
        email = f"{random.choice(names)}{random.randint(1,999)}@{random.choice(domains)}"

    rows.append({
        "Order ID": order_id,
        "Order Date": order_date,
        "Region": region,
        "Sales Rep": rep,
        "Product": product,
        "Quantity": qty,
        "Unit Price (KES)": price,
        "Revenue (KES)": revenue,
        "Order Status": status,
        "Customer Email": email,
    })

df = pd.DataFrame(rows)

# Inject duplicates (full row duplicates)
dupes = df.sample(15, random_state=1)
df = pd.concat([df, dupes], ignore_index=True).sample(frac=1, random_state=42).reset_index(drop=True)

# Save to Excel with some formatting chaos (mimicking real export)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales Data"

# Write a fake title row (common in exported reports)
ws.merge_cells("A1:J1")
ws["A1"] = "SALES & ORDER REPORT — EXPORT — Q1-Q2 2023/2024"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].fill = PatternFill("solid", start_color="4472C4", fgColor="4472C4")
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")

# Write a blank row
ws.append([])

# Write headers on row 3
headers = list(df.columns)
ws.append(headers)
for cell in ws[3]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", start_color="D9E1F2", fgColor="D9E1F2")

# Write data
for _, row in df.iterrows():
    ws.append(list(row))

# Add a totally unrelated notes section at the bottom
ws.append([])
ws.append(["NOTE: This report was auto-generated. Please verify figures before use."])
ws.append(["Last updated by: Finance Team", None, None, None, None, None, None, None, None, "Confidential"])

path = "/home/claude/data_cleaning_pipeline/data/messy_sales_data.xlsx"
wb.save(path)
print(f"Messy dataset saved to: {path}")
print(f"Total rows (including header chaos): {df.shape[0]} data rows")
