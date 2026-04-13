"""
pipeline.py
===========
Advanced Data Cleaning Pipeline — Portfolio Project
Author : Kisavi Shadrack
Email  : shadrackkisavi4@gmail.com

Architecture
------------
  DataLoader        → reads & normalises the raw Excel file
  DataValidator     → runs quality checks BEFORE cleaning, records issues
  DataCleaner       → applies all cleaning transformations
  ReportGenerator   → builds the Summary + Validation Excel reports
  ChartGenerator    → produces PNG visualisations
  Pipeline          → orchestrates all stages, drives logging

Usage
-----
  python pipeline.py                          # uses config/config.yaml
  python pipeline.py --config my_config.yaml  # custom config
"""

import argparse
import logging
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns
import yaml
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# Logging helpers
# ──────────────────────────────────────────────────────────────────────────────

def setup_logging(log_file: str, level: str) -> logging.Logger:
    Path(log_file).parent.mkdir(parents=True, exist_ok=True)
    fmt = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format=fmt,
        handlers=[
            logging.FileHandler(log_file, mode="w"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger("pipeline")


# ──────────────────────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class ValidationIssue:
    row_index: int
    column: str
    issue_type: str
    original_value: Any
    description: str


@dataclass
class PipelineStats:
    raw_rows: int = 0
    clean_rows: int = 0
    duplicates_removed: int = 0
    nulls_fixed: int = 0
    dates_fixed: int = 0
    format_fixes: int = 0
    outliers_flagged: int = 0
    revenue_mismatches: int = 0
    validation_issues: list = field(default_factory=list)


# ──────────────────────────────────────────────────────────────────────────────
# Stage 1 — DataLoader
# ──────────────────────────────────────────────────────────────────────────────

class DataLoader:
    """Reads the raw Excel file, strips the title row and footer noise."""

    def __init__(self, cfg: dict, logger: logging.Logger):
        self.cfg = cfg
        self.log = logger.getChild("DataLoader")

    def load(self) -> pd.DataFrame:
        path = self.cfg["input"]["file_path"]
        sheet = self.cfg["input"]["sheet_name"]
        header_row = self.cfg["input"]["header_row"]
        skip_footer = self.cfg["input"]["skip_footer_rows"]

        self.log.info(f"Loading file: {path}  |  sheet: {sheet}")

        df = pd.read_excel(
            path,
            sheet_name=sheet,
            header=header_row,
            dtype=str,          # read everything as string first — safest
        )

        # Drop the trailing footer rows
        df = df.iloc[: len(df) - skip_footer]

        # Replace all null-marker strings with np.nan
        null_markers = self.cfg["cleaning"]["null_markers"]
        df.replace(null_markers, np.nan, inplace=True)

        # Drop completely empty rows
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)

        self.log.info(f"Loaded {len(df)} rows × {len(df.columns)} columns")
        return df


# ──────────────────────────────────────────────────────────────────────────────
# Stage 2 — DataValidator  (pre-clean snapshot)
# ──────────────────────────────────────────────────────────────────────────────

class DataValidator:
    """
    Records all data quality issues found in the RAW data.
    Results feed the Validation Report. Does NOT modify the DataFrame.
    """

    def __init__(self, cfg: dict, logger: logging.Logger):
        self.cfg = cfg
        self.log = logger.getChild("DataValidator")
        self.issues: list[ValidationIssue] = []

    def _add(self, row, col, itype, val, desc):
        self.issues.append(ValidationIssue(row, col, itype, val, desc))

    def validate(self, df: pd.DataFrame) -> list[ValidationIssue]:
        self.log.info("Running pre-clean validation …")
        cols = self.cfg["columns"]

        for i, row in df.iterrows():

            # Order ID
            oid = row.get(cols["order_id"])
            if pd.isna(oid):
                self._add(i, cols["order_id"], "MISSING", oid, "Order ID is null")
            elif not re.match(self.cfg["cleaning"]["order_id_pattern"], str(oid).strip()):
                self._add(i, cols["order_id"], "FORMAT", oid, f"Order ID does not match expected pattern")

            # Order Date
            dt = row.get(cols["order_date"])
            if pd.isna(dt):
                self._add(i, cols["order_date"], "MISSING", dt, "Order Date is null")
            else:
                parsed = self._try_parse_date(str(dt))
                if parsed is None:
                    self._add(i, cols["order_date"], "INVALID_DATE", dt, f"Cannot parse date: '{dt}'")

            # Region
            region = row.get(cols["region"])
            if pd.isna(region):
                self._add(i, cols["region"], "MISSING", region, "Region is null")
            elif region.strip().title() not in self.cfg["cleaning"]["valid_regions"]:
                self._add(i, cols["region"], "INVALID_CATEGORY", region, f"Unrecognised region")

            # Sales Rep
            rep = row.get(cols["sales_rep"])
            if pd.isna(rep):
                self._add(i, cols["sales_rep"], "MISSING", rep, "Sales Rep is null")

            # Product
            product = row.get(cols["product"])
            if pd.isna(product):
                self._add(i, cols["product"], "MISSING", product, "Product is null")
            elif product.strip().title() not in self.cfg["cleaning"]["valid_products"]:
                self._add(i, cols["product"], "INVALID_CATEGORY", product, f"Unrecognised product")

            # Quantity
            qty = row.get(cols["quantity"])
            if pd.isna(qty):
                self._add(i, cols["quantity"], "MISSING", qty, "Quantity is null")
            else:
                try:
                    q = float(qty)
                    if q < self.cfg["cleaning"]["quantity_min"] or q > self.cfg["cleaning"]["quantity_max"]:
                        self._add(i, cols["quantity"], "OUT_OF_RANGE", qty, f"Quantity {q} out of valid range")
                except ValueError:
                    self._add(i, cols["quantity"], "INVALID_TYPE", qty, f"Non-numeric quantity: '{qty}'")

            # Unit Price
            price = row.get(cols["unit_price"])
            if pd.isna(price):
                self._add(i, cols["unit_price"], "MISSING", price, "Unit Price is null")
            else:
                try:
                    p = float(price)
                    if p < self.cfg["cleaning"]["price_min"] or p > self.cfg["cleaning"]["price_max"]:
                        self._add(i, cols["unit_price"], "OUT_OF_RANGE", price, f"Price {p} out of valid range")
                except ValueError:
                    self._add(i, cols["unit_price"], "INVALID_TYPE", price, f"Non-numeric price: '{price}'")

            # Email
            email = row.get(cols["email"])
            if not pd.isna(email):
                if not re.match(self.cfg["cleaning"]["email_pattern"], str(email).strip()):
                    self._add(i, cols["email"], "INVALID_FORMAT", email, f"Malformed email")

            # Status
            status = row.get(cols["status"])
            if pd.isna(status):
                self._add(i, cols["status"], "MISSING", status, "Order Status is null")
            elif str(status).strip().title() not in self.cfg["cleaning"]["valid_statuses"]:
                self._add(i, cols["status"], "INVALID_CATEGORY", status, f"Unrecognised status")

        self.log.info(f"Validation complete → {len(self.issues)} issues found")
        return self.issues

    def _try_parse_date(self, s: str):
        for fmt in self.cfg["cleaning"]["date_formats"]:
            try:
                return datetime.strptime(s.strip(), fmt)
            except:
                pass
        return None


# ──────────────────────────────────────────────────────────────────────────────
# Stage 3 — DataCleaner
# ──────────────────────────────────────────────────────────────────────────────

class DataCleaner:
    """Applies all transformations. Returns a clean DataFrame + stats."""

    def __init__(self, cfg: dict, logger: logging.Logger):
        self.cfg = cfg
        self.log = logger.getChild("DataCleaner")

    def clean(self, df: pd.DataFrame, stats: PipelineStats) -> pd.DataFrame:
        self.log.info("Starting cleaning …")
        stats.raw_rows = len(df)
        cols = self.cfg["columns"]

        df = df.copy()

        # ── 1. Remove duplicates ────────────────────────────────────────────
        before = len(df)
        df.drop_duplicates(inplace=True)
        df.reset_index(drop=True, inplace=True)
        stats.duplicates_removed = before - len(df)
        self.log.info(f"  Duplicates removed: {stats.duplicates_removed}")

        # ── 2. Standardise text columns ────────────────────────────────────
        text_cols = [cols["region"], cols["sales_rep"], cols["product"],
                     cols["status"], cols["order_id"]]
        for c in text_cols:
            if c in df.columns:
                # strip whitespace and collapse internal spaces
                df[c] = df[c].astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
                df[c] = df[c].replace("nan", np.nan)

        # Title-case region, product, status
        for c in [cols["region"], cols["product"], cols["status"]]:
            if c in df.columns:
                df[c] = df[c].str.title()

        # Proper-case sales rep names
        df[cols["sales_rep"]] = df[cols["sales_rep"]].str.title()

        stats.format_fixes += 1
        self.log.info("  Text standardisation done")

        # ── 3. Fix Order IDs ───────────────────────────────────────────────
        pattern = self.cfg["cleaning"]["order_id_pattern"]
        def fix_order_id(v):
            if pd.isna(v):
                return np.nan
            v = str(v).strip()
            if re.match(pattern, v):
                return v
            # Try to rescue e.g. "ORD0023" → "ORD-0023"
            m = re.match(r"ORD(\d{4})", v)
            if m:
                return f"ORD-{m.group(1)}"
            return np.nan
        df[cols["order_id"]] = df[cols["order_id"]].apply(fix_order_id)

        # ── 4. Parse dates ─────────────────────────────────────────────────
        def parse_date(v):
            if pd.isna(v):
                return pd.NaT
            for fmt in self.cfg["cleaning"]["date_formats"]:
                try:
                    return datetime.strptime(str(v).strip(), fmt)
                except:
                    pass
            return pd.NaT

        raw_dates = df[cols["order_date"]].copy()
        df[cols["order_date"]] = df[cols["order_date"]].apply(parse_date)
        stats.dates_fixed = int((raw_dates.notna() & df[cols["order_date"]].isna()).sum())
        self.log.info(f"  Unparseable dates set to NaT: {stats.dates_fixed}")

        # ── 5. Numeric coercion ────────────────────────────────────────────
        for c in [cols["quantity"], cols["unit_price"], cols["revenue"]]:
            df[c] = pd.to_numeric(df[c], errors="coerce")

        # ── 6. Range clamp / nullify ───────────────────────────────────────
        q_min, q_max = self.cfg["cleaning"]["quantity_min"], self.cfg["cleaning"]["quantity_max"]
        p_min, p_max = self.cfg["cleaning"]["price_min"], self.cfg["cleaning"]["price_max"]

        df.loc[~df[cols["quantity"]].between(q_min, q_max), cols["quantity"]] = np.nan
        df.loc[~df[cols["unit_price"]].between(p_min, p_max), cols["unit_price"]] = np.nan

        # ── 7. Revenue reconciliation ──────────────────────────────────────
        tol = self.cfg["cleaning"]["revenue_tolerance_pct"]
        calc_rev = df[cols["quantity"]] * df[cols["unit_price"]]
        mismatch_mask = (
            df[cols["revenue"]].notna() &
            calc_rev.notna() &
            ((df[cols["revenue"]] - calc_rev).abs() / calc_rev.clip(lower=1) > tol)
        )
        stats.revenue_mismatches = int(mismatch_mask.sum())
        # Trust qty × price; overwrite bad revenue
        df.loc[mismatch_mask, cols["revenue"]] = calc_rev[mismatch_mask].round(2)
        # Fill missing revenue where qty & price available
        missing_rev = df[cols["revenue"]].isna() & calc_rev.notna()
        df.loc[missing_rev, cols["revenue"]] = calc_rev[missing_rev].round(2)
        self.log.info(f"  Revenue mismatches corrected: {stats.revenue_mismatches}")

        # ── 8. Outlier detection (Z-score) ─────────────────────────────────
        z_thresh = self.cfg["cleaning"]["outlier_zscore_threshold"]
        for c in [cols["unit_price"], cols["revenue"]]:
            col_data = df[c].dropna()
            if len(col_data) > 10:
                z = (col_data - col_data.mean()) / col_data.std()
                outlier_idx = z[z.abs() > z_thresh].index
                df.loc[outlier_idx, f"{c}_outlier_flag"] = True
                stats.outliers_flagged += len(outlier_idx)
        self.log.info(f"  Outliers flagged: {stats.outliers_flagged}")

        # ── 9. Email validation ────────────────────────────────────────────
        email_pat = self.cfg["cleaning"]["email_pattern"]
        invalid_email = df[cols["email"]].notna() & ~df[cols["email"]].str.match(email_pat, na=False)
        df.loc[invalid_email, cols["email"]] = np.nan

        # ── 10. Add derived columns ────────────────────────────────────────
        df["Year"] = df[cols["order_date"]].dt.year
        df["Month"] = df[cols["order_date"]].dt.month
        df["Quarter"] = df[cols["order_date"]].dt.to_period("Q").astype(str)

        # ── 11. Track remaining nulls ──────────────────────────────────────
        stats.nulls_fixed = int(df.isna().sum().sum())

        stats.clean_rows = len(df)
        self.log.info(f"Cleaning done. Clean rows: {stats.clean_rows}")
        return df


# ──────────────────────────────────────────────────────────────────────────────
# Stage 4 — ReportGenerator
# ──────────────────────────────────────────────────────────────────────────────

class ReportGenerator:
    """Writes formatted Excel summary and validation reports."""

    BLUE   = "2F5496"
    LBLUE  = "D9E1F2"
    GREEN  = "375623"
    LGREEN = "E2EFDA"
    RED    = "C00000"
    LRED   = "FFE0E0"
    ORANGE = "F4B942"

    def __init__(self, cfg: dict, logger: logging.Logger):
        self.cfg = cfg
        self.log = logger.getChild("ReportGenerator")

    # ── helpers ──────────────────────────────────────────────────────────────

    def _hdr_style(self, cell, bg=None, fg="FFFFFF", bold=True, size=11):
        cell.font = Font(bold=bold, color=fg, size=size)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _border_all(self, ws, min_row, min_col, max_row, max_col):
        thin = Side(style="thin")
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                 min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Summary Report ───────────────────────────────────────────────────────

    def write_summary_report(self, df: pd.DataFrame, stats: PipelineStats):
        path = self.cfg["output"]["summary_report"]
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        cols = self.cfg["columns"]

        wb = openpyxl.Workbook()
        self._sheet_pipeline_summary(wb, stats)
        self._sheet_sales_by_region(wb, df, cols)
        self._sheet_sales_by_product(wb, df, cols)
        self._sheet_sales_by_rep(wb, df, cols)
        self._sheet_monthly_trend(wb, df, cols)

        del wb["Sheet"]  # remove default sheet
        wb.save(path)
        self.log.info(f"Summary report saved → {path}")

    def _title_row(self, ws, title, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = ws.cell(1, 1, title)
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=self.BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.append([])  # blank row

    def _sheet_pipeline_summary(self, wb, stats: PipelineStats):
        ws = wb.create_sheet("Pipeline Summary")
        self._title_row(ws, "Pipeline Run Summary", 3)
        headers = ["Metric", "Value", "Notes"]
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            self._hdr_style(ws.cell(3, i), bg=self.BLUE)

        rows = [
            ("Raw Rows Loaded",        stats.raw_rows,              "Before deduplication"),
            ("Clean Rows Output",       stats.clean_rows,            "After all transformations"),
            ("Duplicates Removed",      stats.duplicates_removed,    "Exact row duplicates"),
            ("Unparseable Dates",       stats.dates_fixed,           "Set to NaT"),
            ("Revenue Mismatches Fixed",stats.revenue_mismatches,    "Overwritten with qty × price"),
            ("Outliers Flagged",        stats.outliers_flagged,      "Z-score > 3 (not removed)"),
            ("Remaining Null Values",   stats.nulls_fixed,           "Across all columns"),
            ("Validation Issues Found", len(stats.validation_issues),"See Validation Report"),
            ("Run Timestamp",           datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
        ]
        for i, r in enumerate(rows):
            ws.append(list(r))
            fill = self.LGREEN if i % 2 == 0 else "FFFFFF"
            for j in range(1, 4):
                ws.cell(ws.max_row, j).fill = PatternFill("solid", fgColor=fill)

        self._border_all(ws, 3, 1, ws.max_row, 3)
        self._auto_width(ws)

    def _sheet_sales_by_region(self, wb, df, cols):
        ws = wb.create_sheet("Sales by Region")
        self._title_row(ws, "Revenue & Orders by Region", 4)
        ws.append(["Region", "Total Revenue (KES)", "Orders", "Avg Revenue/Order"])
        for j in range(1, 5):
            self._hdr_style(ws.cell(3, j), bg=self.BLUE)

        grp = df.groupby(cols["region"]).agg(
            Revenue=(cols["revenue"], "sum"),
            Orders=(cols["order_id"], "count"),
        ).reset_index().sort_values("Revenue", ascending=False)

        for _, r in grp.iterrows():
            avg = r["Revenue"] / r["Orders"] if r["Orders"] else 0
            ws.append([r[cols["region"]], round(r["Revenue"], 2), r["Orders"], round(avg, 2)])

        self._border_all(ws, 3, 1, ws.max_row, 4)
        self._auto_width(ws)

    def _sheet_sales_by_product(self, wb, df, cols):
        ws = wb.create_sheet("Sales by Product")
        self._title_row(ws, "Revenue & Quantity by Product", 4)
        ws.append(["Product", "Total Revenue (KES)", "Units Sold", "Avg Unit Price (KES)"])
        for j in range(1, 5):
            self._hdr_style(ws.cell(3, j), bg=self.BLUE)

        grp = df.groupby(cols["product"]).agg(
            Revenue=(cols["revenue"], "sum"),
            Units=(cols["quantity"], "sum"),
            AvgPrice=(cols["unit_price"], "mean"),
        ).reset_index().sort_values("Revenue", ascending=False)

        for _, r in grp.iterrows():
            ws.append([r[cols["product"]], round(r["Revenue"], 2),
                       round(r["Units"], 0), round(r["AvgPrice"], 2)])

        self._border_all(ws, 3, 1, ws.max_row, 4)
        self._auto_width(ws)

    def _sheet_sales_by_rep(self, wb, df, cols):
        ws = wb.create_sheet("Sales by Rep")
        self._title_row(ws, "Performance by Sales Representative", 4)
        ws.append(["Sales Rep", "Total Revenue (KES)", "Orders", "Completed Orders"])
        for j in range(1, 5):
            self._hdr_style(ws.cell(3, j), bg=self.BLUE)

        completed = df[df[cols["status"]] == "Completed"]
        grp = df.groupby(cols["sales_rep"]).agg(
            Revenue=(cols["revenue"], "sum"),
            Orders=(cols["order_id"], "count"),
        ).reset_index()
        comp_grp = completed.groupby(cols["sales_rep"]).size().reset_index(name="Completed")
        grp = grp.merge(comp_grp, on=cols["sales_rep"], how="left").fillna(0)
        grp.sort_values("Revenue", ascending=False, inplace=True)

        for _, r in grp.iterrows():
            ws.append([r[cols["sales_rep"]], round(r["Revenue"], 2),
                       int(r["Orders"]), int(r["Completed"])])

        self._border_all(ws, 3, 1, ws.max_row, 4)
        self._auto_width(ws)

    def _sheet_monthly_trend(self, wb, df, cols):
        ws = wb.create_sheet("Monthly Trend")
        self._title_row(ws, "Monthly Revenue Trend", 3)
        ws.append(["Year-Month", "Total Revenue (KES)", "Orders"])
        for j in range(1, 4):
            self._hdr_style(ws.cell(3, j), bg=self.BLUE)

        df2 = df.dropna(subset=[cols["order_date"]])
        df2 = df2.copy()
        df2["YM"] = df2[cols["order_date"]].dt.to_period("M").astype(str)
        grp = df2.groupby("YM").agg(
            Revenue=(cols["revenue"], "sum"),
            Orders=(cols["order_id"], "count"),
        ).reset_index().sort_values("YM")

        for _, r in grp.iterrows():
            ws.append([r["YM"], round(r["Revenue"], 2), int(r["Orders"])])

        self._border_all(ws, 3, 1, ws.max_row, 3)
        self._auto_width(ws)

    # ── Validation Report ────────────────────────────────────────────────────

    def write_validation_report(self, issues: list[ValidationIssue]):
        path = self.cfg["output"]["validation_report"]
        Path(path).parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Issues"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Data Validation Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=self.RED)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26
        ws.append([])

        headers = ["Row Index", "Column", "Issue Type", "Original Value", "Description"]
        ws.append(headers)
        for j, h in enumerate(headers, 1):
            cell = ws.cell(3, j)
            self._hdr_style(cell, bg="C00000")

        type_colors = {
            "MISSING": "FFE0E0",
            "INVALID_DATE": "FFF2CC",
            "INVALID_TYPE": "FFF2CC",
            "OUT_OF_RANGE": "FCE4D6",
            "INVALID_FORMAT": "EAF1FB",
            "INVALID_CATEGORY": "F4CCCC",
            "FORMAT": "EAF1FB",
        }

        for issue in issues:
            ws.append([issue.row_index, issue.column, issue.issue_type,
                       str(issue.original_value), issue.description])
            color = type_colors.get(issue.issue_type, "FFFFFF")
            for j in range(1, 6):
                ws.cell(ws.max_row, j).fill = PatternFill("solid", fgColor=color)

        self._border_all(ws, 3, 1, ws.max_row, 5)

        # Summary sheet
        ws2 = wb.create_sheet("Issue Summary")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "Issues by Type"
        self._hdr_style(ws2["A1"], bg=self.RED, size=12)
        ws2.row_dimensions[1].height = 24
        ws2.append([])
        ws2.append(["Issue Type", "Count", "% of Total"])
        for j in range(1, 4):
            self._hdr_style(ws2.cell(3, j), bg="C00000")

        from collections import Counter
        counts = Counter(i.issue_type for i in issues)
        total = len(issues)
        for itype, cnt in sorted(counts.items(), key=lambda x: -x[1]):
            ws2.append([itype, cnt, f"{cnt/total*100:.1f}%" if total else "0%"])

        self._border_all(ws2, 3, 1, ws2.max_row, 3)
        self._auto_width(ws)
        self._auto_width(ws2)
        wb.save(path)
        self.log.info(f"Validation report saved → {path}")

    # ── Clean file ───────────────────────────────────────────────────────────

    def write_clean_file(self, df: pd.DataFrame):
        path = self.cfg["output"]["clean_file"]
        Path(path).parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clean Sales Data"

        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        ws["A1"] = f"Clean Sales Data — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="375623")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24
        ws.append([])

        # Headers
        headers = list(df.columns)
        ws.append(headers)
        for j, h in enumerate(headers, 1):
            cell = ws.cell(3, j)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="375623")
            cell.alignment = Alignment(horizontal="center")

        # Data rows with alternating fill
        for i, (_, row) in enumerate(df.iterrows()):
            vals = []
            for v in row:
                if isinstance(v, pd.Timestamp):
                    vals.append(v.strftime("%Y-%m-%d") if not pd.isna(v) else "")
                elif isinstance(v, float) and np.isnan(v):
                    vals.append("")
                else:
                    vals.append(v)
            ws.append(vals)
            fill_color = "E2EFDA" if i % 2 == 0 else "FFFFFF"
            for j in range(1, len(headers) + 1):
                ws.cell(ws.max_row, j).fill = PatternFill("solid", fgColor=fill_color)

        self._border_all(ws, 3, 1, ws.max_row, len(headers))
        self._auto_width(ws)
        wb.save(path)
        self.log.info(f"Clean file saved → {path}")


# ──────────────────────────────────────────────────────────────────────────────
# Stage 5 — ChartGenerator
# ──────────────────────────────────────────────────────────────────────────────

class ChartGenerator:
    """Produces publication-quality PNG charts."""

    PALETTE = ["#2F5496", "#375623", "#C00000", "#F4B942", "#7030A0", "#00B0F0"]

    def __init__(self, cfg: dict, logger: logging.Logger):
        self.cfg = cfg
        self.log = logger.getChild("ChartGenerator")
        self.charts_dir = Path(cfg["output"]["charts_dir"])
        self.charts_dir.mkdir(parents=True, exist_ok=True)
        plt.rcParams.update({
            "font.family": "DejaVu Sans",
            "axes.spines.top": False,
            "axes.spines.right": False,
            "figure.dpi": 150,
        })

    def generate_all(self, df: pd.DataFrame):
        cols = self.cfg["columns"]
        self.log.info("Generating charts …")
        self._revenue_by_region(df, cols)
        self._revenue_by_product(df, cols)
        self._monthly_trend(df, cols)
        self._status_distribution(df, cols)
        self._price_distribution(df, cols)
        self.log.info(f"Charts saved to {self.charts_dir}")

    def _save(self, fig, name):
        p = self.charts_dir / name
        fig.savefig(p, bbox_inches="tight")
        plt.close(fig)

    def _revenue_by_region(self, df, cols):
        grp = df.groupby(cols["region"])[cols["revenue"]].sum().sort_values(ascending=True)
        fig, ax = plt.subplots(figsize=(8, 4))
        bars = ax.barh(grp.index, grp.values / 1e6, color=self.PALETTE[:len(grp)])
        ax.set_xlabel("Revenue (KES Millions)")
        ax.set_title("Total Revenue by Region", fontsize=13, fontweight="bold")
        for bar in bars:
            ax.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                    f"{bar.get_width():.1f}M", va="center", fontsize=9)
        self._save(fig, "01_revenue_by_region.png")

    def _revenue_by_product(self, df, cols):
        grp = df.groupby(cols["product"])[cols["revenue"]].sum().sort_values(ascending=False)
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.bar(grp.index, grp.values / 1e6, color=self.PALETTE[:len(grp)], edgecolor="white")
        ax.set_ylabel("Revenue (KES Millions)")
        ax.set_title("Revenue by Product Category", fontsize=13, fontweight="bold")
        ax.tick_params(axis="x", rotation=15)
        self._save(fig, "02_revenue_by_product.png")

    def _monthly_trend(self, df, cols):
        df2 = df.dropna(subset=[cols["order_date"]]).copy()
        df2["YM"] = df2[cols["order_date"]].dt.to_period("M")
        grp = df2.groupby("YM")[cols["revenue"]].sum().reset_index()
        grp["YM_str"] = grp["YM"].astype(str)
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.plot(grp["YM_str"], grp[cols["revenue"]] / 1e6,
                marker="o", color=self.PALETTE[0], linewidth=2, markersize=5)
        ax.fill_between(grp["YM_str"], grp[cols["revenue"]] / 1e6, alpha=0.15, color=self.PALETTE[0])
        ax.set_ylabel("Revenue (KES Millions)")
        ax.set_title("Monthly Revenue Trend", fontsize=13, fontweight="bold")
        ax.tick_params(axis="x", rotation=45)
        plt.tight_layout()
        self._save(fig, "03_monthly_trend.png")

    def _status_distribution(self, df, cols):
        grp = df[cols["status"]].value_counts()
        fig, ax = plt.subplots(figsize=(6, 6))
        wedges, texts, autotexts = ax.pie(
            grp.values, labels=grp.index,
            autopct="%1.1f%%", colors=self.PALETTE[:len(grp)],
            startangle=140, pctdistance=0.8,
        )
        for t in autotexts:
            t.set_fontsize(10)
        ax.set_title("Order Status Distribution", fontsize=13, fontweight="bold")
        self._save(fig, "04_status_distribution.png")

    def _price_distribution(self, df, cols):
        data = df[cols["unit_price"]].dropna()
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.hist(data / 1000, bins=30, color=self.PALETTE[0], edgecolor="white")
        ax.set_xlabel("Unit Price (KES Thousands)")
        ax.set_ylabel("Frequency")
        ax.set_title("Unit Price Distribution", fontsize=13, fontweight="bold")
        median = data.median() / 1000
        ax.axvline(median, color=self.PALETTE[2], linestyle="--", linewidth=1.5, label=f"Median: {median:.1f}K")
        ax.legend()
        self._save(fig, "05_price_distribution.png")


# ──────────────────────────────────────────────────────────────────────────────
# Orchestrator — Pipeline
# ──────────────────────────────────────────────────────────────────────────────

class Pipeline:
    def __init__(self, config_path: str):
        with open(config_path) as f:
            self.cfg = yaml.safe_load(f)
        self.logger = setup_logging(
            self.cfg["logging"]["log_file"],
            self.cfg["logging"]["level"],
        )
        self.stats = PipelineStats()

    def run(self):
        log = self.logger
        log.info("=" * 60)
        log.info(f"  {self.cfg['pipeline']['name']}  v{self.cfg['pipeline']['version']}")
        log.info(f"  Author: {self.cfg['pipeline']['author']}")
        log.info("=" * 60)

        # Stage 1 — Load
        loader = DataLoader(self.cfg, log)
        raw_df = loader.load()

        # Stage 2 — Validate (on raw data)
        validator = DataValidator(self.cfg, log)
        issues = validator.validate(raw_df)
        self.stats.validation_issues = issues

        # Stage 3 — Clean
        cleaner = DataCleaner(self.cfg, log)
        clean_df = cleaner.clean(raw_df, self.stats)

        # Stage 4 — Reports
        reporter = ReportGenerator(self.cfg, log)
        reporter.write_clean_file(clean_df)
        reporter.write_summary_report(clean_df, self.stats)
        reporter.write_validation_report(issues)

        # Stage 5 — Charts
        charter = ChartGenerator(self.cfg, log)
        charter.generate_all(clean_df)

        log.info("=" * 60)
        log.info("Pipeline complete. All outputs written.")
        log.info(f"  Raw rows     : {self.stats.raw_rows}")
        log.info(f"  Clean rows   : {self.stats.clean_rows}")
        log.info(f"  Duplicates   : {self.stats.duplicates_removed}")
        log.info(f"  Outliers     : {self.stats.outliers_flagged}")
        log.info(f"  Val. issues  : {len(self.stats.validation_issues)}")
        log.info("=" * 60)


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel Data Cleaning Pipeline")
    parser.add_argument("--config", default="config/config.yaml", help="Path to YAML config file")
    args = parser.parse_args()
    Pipeline(args.config).run()
