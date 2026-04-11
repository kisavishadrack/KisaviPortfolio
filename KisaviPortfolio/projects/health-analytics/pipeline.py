"""
pipeline.py
===========
Health Analytics Pipeline — PostgreSQL + Dashboard
Author : Kisavi Shadrack | shadrackkisavi4@gmail.com

Stages
------
  1. DataLoader       — reads all four CSVs
  2. DataCleaner      — resolves 20+ issue categories per table
  3. DataValidator    — generates DQ report
  4. SQLExporter      — writes INSERT SQL + view definitions
  5. AnalyticsEngine  — runs 12 analytical queries, produces KPI summary
  6. ReportGenerator  — writes Excel validation + summary reports
  7. ChartGenerator   — 6 PNG charts for the HTML dashboard

Usage
-----
  python pipeline.py                 # SQLite simulation (no PG required)
  python pipeline.py --pg            # write SQL files for PostgreSQL
"""

import argparse, logging, os, re, sys, warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import sqlite3
import yaml
from collections import Counter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Logging ───────────────────────────────────────────────────────────────────

def setup_logging(log_file, level="INFO"):
    Path(log_file).parent.mkdir(parents=True, exist_ok=True)
    fmt = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    logging.basicConfig(level=getattr(logging, level),
        format=fmt,
        handlers=[logging.FileHandler(log_file, "w"), logging.StreamHandler(sys.stdout)])
    return logging.getLogger("health_pipeline")

# ── Config ────────────────────────────────────────────────────────────────────

CONFIG = {
    "null_markers": ["N/A","n/a","NA","","None","NaN","TBD","Unknown","UNKNOWN","unknown",
                     "pending","error","N/a","not provided"],
    "date_formats": ["%d/%m/%Y","%Y-%m-%d","%m-%d-%Y","%d-%m-%Y","%m/%d/%Y"],
    "valid_genders": ["Male","Female","Other"],
    "valid_blood_groups": ["A+","A-","B+","B-","O+","O-","AB+","AB-"],
    "valid_wards": ["General","ICU","Maternity","Paediatrics","Surgical","Orthopaedics","Oncology"],
    "valid_outcomes": ["Recovered","Referred","Deceased","Absconded","Against Medical Advice"],
    "valid_result_statuses": ["Normal","Abnormal","Critical","Borderline"],
    "valid_opd_outcomes": ["Prescription given","Referred to specialist","Admitted",
                           "Investigations ordered","Counselled & discharged","Follow-up scheduled"],
    "cost_min": 100, "cost_max": 500000,
    "wait_min": 1, "wait_max": 600,
    "fee_min": 50, "fee_max": 10000,
}

# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class DQIssue:
    table: str; row: int; column: str; issue_type: str
    original_value: Any; description: str

@dataclass
class PipelineStats:
    tables: dict = field(default_factory=dict)
    dq_issues: list = field(default_factory=list)

# ── 1. DataLoader ─────────────────────────────────────────────────────────────

class DataLoader:
    def __init__(self, log): self.log = log.getChild("Loader")

    def load_all(self):
        files = {
            "patients":           "data/patients.csv",
            "admissions":         "data/admissions.csv",
            "lab_tests":          "data/lab_tests.csv",
            "outpatient_visits":  "data/outpatient_visits.csv",
        }
        dfs = {}
        for name, path in files.items():
            df = pd.read_csv(path, dtype=str)
            df.replace(CONFIG["null_markers"], np.nan, inplace=True)
            df.dropna(how="all", inplace=True)
            df.reset_index(drop=True, inplace=True)
            self.log.info(f"Loaded {name}: {len(df)} rows × {len(df.columns)} cols")
            dfs[name] = df
        return dfs

# ── 2. DataCleaner ────────────────────────────────────────────────────────────

class DataCleaner:
    def __init__(self, log, stats): self.log = log.getChild("Cleaner"); self.stats = stats

    def _parse_date(self, s):
        if pd.isna(s): return pd.NaT
        for fmt in CONFIG["date_formats"]:
            try: return pd.to_datetime(str(s).strip(), format=fmt)
            except: pass
        return pd.NaT

    def _std_text(self, series, title=True):
        s = series.astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
        s = s.replace("nan", np.nan)
        return s.str.title() if title else s

    def _numeric(self, series, lo=None, hi=None):
        n = pd.to_numeric(series, errors="coerce")
        if lo is not None: n = n.where(n >= lo, np.nan)
        if hi is not None: n = n.where(n <= hi, np.nan)
        return n

    def _map_values(self, series, valid_list, extra_map=None):
        """Title-case + apply extra synonyms map, invalid → NaN."""
        mapped = self._std_text(series, title=True)
        if extra_map:
            mapped = mapped.map(lambda v: extra_map.get(v, v) if pd.notna(v) else v)
        return mapped.where(mapped.isin(valid_list + [np.nan]), np.nan)

    def clean_patients(self, df):
        self.log.info("Cleaning patients …")
        raw = len(df)
        df = df.copy()
        df.drop_duplicates(inplace=True)

        # IDs: rescue PAT0001 → PAT-0001
        def fix_pid(v):
            if pd.isna(v): return np.nan
            v = str(v).strip()
            if re.match(r"^PAT-\d{4}$", v): return v
            m = re.match(r"^P(?:AT)?-?(\d{4})$", v)
            return f"PAT-{m.group(1)}" if m else np.nan
        df["patient_id"] = df["patient_id"].apply(fix_pid)
        df.dropna(subset=["patient_id"], inplace=True)
        df.drop_duplicates(subset=["patient_id"], keep="first", inplace=True)

        df["first_name"] = self._std_text(df["first_name"], title=True)
        df["last_name"]  = self._std_text(df["last_name"],  title=True)
        df.dropna(subset=["first_name","last_name"], inplace=True)

        df["date_of_birth"] = df["date_of_birth"].apply(self._parse_date)
        today = pd.Timestamp("2024-06-30")
        df["age_years"] = ((today - df["date_of_birth"]).dt.days / 365.25).apply(
            lambda x: int(x) if pd.notna(x) and 0 < x < 120 else np.nan)

        gender_map = {"M":"Male","F":"Female","Male":"Male","Female":"Female",
                      "MALE":"Male","FEMALE":"Female","male":"Male","female":"Female"}
        df["gender"] = df["gender"].map(lambda v: gender_map.get(str(v).strip(), np.nan) if pd.notna(v) else np.nan)

        bg_map = {g: g for g in CONFIG["valid_blood_groups"]}
        df["blood_group"] = df["blood_group"].map(lambda v: bg_map.get(str(v).strip().upper(), np.nan) if pd.notna(v) else np.nan)

        df["county"] = self._std_text(df["county"], title=True)
        df["registration_date"] = df["registration_date"].apply(self._parse_date)
        df["is_active"] = df["is_active"].map(
            lambda v: True if str(v).strip() in ["1","1.0","TRUE","True","yes"] else
                      False if str(v).strip() in ["0","0.0","FALSE","False","no"] else np.nan)

        df.reset_index(drop=True, inplace=True)
        self.log.info(f"  patients: {raw} → {len(df)} rows")
        self.stats.tables["patients"] = {"raw": raw, "clean": len(df)}
        return df

    def clean_admissions(self, df, valid_pids):
        self.log.info("Cleaning admissions …")
        raw = len(df)
        df = df.copy()
        df.drop_duplicates(inplace=True)

        def fix_aid(v):
            if pd.isna(v): return np.nan
            v = str(v).strip()
            return v if re.match(r"^ADM-\d{5}$", v) else np.nan
        df["admission_id"] = df["admission_id"].apply(fix_aid)
        df.dropna(subset=["admission_id"], inplace=True)
        df.drop_duplicates(subset=["admission_id"], keep="first", inplace=True)

        def fix_pid(v):
            if pd.isna(v): return np.nan
            v = str(v).strip()
            m = re.match(r"^PAT-?(\d{4})$", v)
            return f"PAT-{m.group(1)}" if m else np.nan
        df["patient_id"] = df["patient_id"].apply(fix_pid)
        df = df[df["patient_id"].isin(valid_pids)]

        df["admission_date"]  = df["admission_date"].apply(self._parse_date)
        df["discharge_date"]  = df["discharge_date"].apply(self._parse_date)
        df.dropna(subset=["admission_date"], inplace=True)

        # Fix discharge before admission
        bad_dis = df["discharge_date"] < df["admission_date"]
        df.loc[bad_dis, "discharge_date"] = np.nan
        self.log.info(f"  Discharge-before-admission fixed: {bad_dis.sum()}")

        df["length_of_stay"] = (df["discharge_date"] - df["admission_date"]).dt.days
        df["length_of_stay"] = df["length_of_stay"].where(df["length_of_stay"] >= 0, np.nan)

        df["ward"] = self._map_values(df["ward"], CONFIG["valid_wards"])
        df["diagnosis"] = self._std_text(df["diagnosis"], title=True)

        outcome_map = {"Recovered":"Recovered","Dead":"Deceased","Deceased":"Deceased",
                       "Referred":"Referred","Ref":"Referred","Absconded":"Absconded",
                       "Against Medical Advice":"Against Medical Advice","Ama":"Against Medical Advice",
                       "Alive":"Recovered"}
        df["discharge_outcome"] = df["discharge_outcome"].map(
            lambda v: outcome_map.get(str(v).strip().title(), np.nan) if pd.notna(v) else np.nan)

        df["total_cost_kes"] = self._numeric(df["total_cost_kes"],
            lo=CONFIG["cost_min"], hi=CONFIG["cost_max"])
        df["bed_number"] = pd.to_numeric(df["bed_number"], errors="coerce").astype("Int64")
        df["attending_doctor"] = self._std_text(df["attending_doctor"], title=True)

        df.reset_index(drop=True, inplace=True)
        self.log.info(f"  admissions: {raw} → {len(df)} rows")
        self.stats.tables["admissions"] = {"raw": raw, "clean": len(df)}
        return df

    def clean_lab_tests(self, df, valid_pids):
        self.log.info("Cleaning lab tests …")
        raw = len(df)
        df = df.copy()
        df.drop_duplicates(inplace=True)

        df["lab_id"] = df["lab_id"].apply(
            lambda v: v if pd.notna(v) and re.match(r"^LAB-\d{6}$", str(v).strip()) else np.nan)
        df.dropna(subset=["lab_id"], inplace=True)
        df.drop_duplicates(subset=["lab_id"], keep="first", inplace=True)

        def fix_pid(v):
            if pd.isna(v): return np.nan
            m = re.match(r"^PAT-?(\d{4})$", str(v).strip())
            return f"PAT-{m.group(1)}" if m else np.nan
        df["patient_id"] = df["patient_id"].apply(fix_pid)
        df = df[df["patient_id"].isin(valid_pids)]

        df["ordered_date"] = df["ordered_date"].apply(self._parse_date)
        df["result_date"]  = df["result_date"].apply(self._parse_date)
        df.dropna(subset=["ordered_date"], inplace=True)

        bad = df["result_date"] < df["ordered_date"]
        df.loc[bad, "result_date"] = np.nan
        self.log.info(f"  Result-before-order fixed: {bad.sum()}")

        df["turnaround_hours"] = (
            (df["result_date"] - df["ordered_date"]).dt.total_seconds() / 3600
        ).round(2)
        df["turnaround_hours"] = df["turnaround_hours"].where(df["turnaround_hours"] > 0, np.nan)

        df["test_name"] = self._std_text(df["test_name"], title=True)
        df["result_value"] = pd.to_numeric(df["result_value"], errors="coerce")
        df.loc[df["result_value"] < 0, "result_value"] = np.nan

        status_map = {"Normal":"Normal","Abnormal":"Abnormal","Crit":"Critical",
                      "Critical":"Critical","Borderline":"Borderline","ABNORMAL":"Abnormal"}
        df["result_status"] = df["result_status"].map(
            lambda v: status_map.get(str(v).strip().title(), np.nan) if pd.notna(v) else np.nan)
        df["is_abnormal"] = df["result_status"].isin(["Abnormal","Critical"])

        df.reset_index(drop=True, inplace=True)
        self.log.info(f"  lab_tests: {raw} → {len(df)} rows")
        self.stats.tables["lab_tests"] = {"raw": raw, "clean": len(df)}
        return df

    def clean_outpatient(self, df, valid_pids):
        self.log.info("Cleaning outpatient visits …")
        raw = len(df)
        df = df.copy()
        df.drop_duplicates(inplace=True)

        df["visit_id"] = df["visit_id"].apply(
            lambda v: v if pd.notna(v) and re.match(r"^OPD-\d{6}$", str(v).strip()) else np.nan)
        df.dropna(subset=["visit_id"], inplace=True)
        df.drop_duplicates(subset=["visit_id"], keep="first", inplace=True)

        def fix_pid(v):
            if pd.isna(v): return np.nan
            m = re.match(r"^PAT-?(\d{4})$", str(v).strip())
            return f"PAT-{m.group(1)}" if m else np.nan
        df["patient_id"] = df["patient_id"].apply(fix_pid)
        df = df[df["patient_id"].isin(valid_pids)]

        df["visit_date"]    = df["visit_date"].apply(self._parse_date)
        df["follow_up_date"]= df["follow_up_date"].apply(self._parse_date)
        df.dropna(subset=["visit_date"], inplace=True)

        df["clinic"] = self._std_text(df["clinic"], title=True)
        df["chief_complaint"] = self._std_text(df["chief_complaint"], title=True)
        df["attending_doctor"] = self._std_text(df["attending_doctor"], title=True)

        outcome_map = {"Prescription Given":"Prescription given",
                       "Referred To Specialist":"Referred to specialist",
                       "Admitted":"Admitted","Investigations Ordered":"Investigations ordered",
                       "Counselled & Discharged":"Counselled & discharged",
                       "Follow-Up Scheduled":"Follow-up scheduled",
                       "Given Rx":"Prescription given","Referred":"Referred to specialist",
                       "Discharged":"Counselled & discharged"}
        df["outcome"] = df["outcome"].map(
            lambda v: outcome_map.get(str(v).strip().title(), np.nan) if pd.notna(v) else np.nan)

        df["wait_time_mins"] = self._numeric(df["wait_time_mins"],
            lo=CONFIG["wait_min"], hi=CONFIG["wait_max"]).astype("Int64")
        df["consultation_fee_kes"] = self._numeric(df["consultation_fee_kes"],
            lo=CONFIG["fee_min"], hi=CONFIG["fee_max"])

        df.reset_index(drop=True, inplace=True)
        self.log.info(f"  outpatient_visits: {raw} → {len(df)} rows")
        self.stats.tables["outpatient_visits"] = {"raw": raw, "clean": len(df)}
        return df

# ── 3. AnalyticsEngine ────────────────────────────────────────────────────────

class AnalyticsEngine:
    """Runs all analytical queries using SQLite as a local PG simulation."""

    def __init__(self, log): self.log = log.getChild("Analytics")

    def load_to_sqlite(self, dfs):
        conn = sqlite3.connect(":memory:")
        for name, df in dfs.items():
            df.to_sql(name, conn, if_exists="replace", index=False)
        self.log.info("Data loaded into in-memory SQLite")
        return conn

    def run(self, conn):
        queries = {
            "admissions_by_ward": """
                SELECT ward, COUNT(*) AS total_admissions,
                       ROUND(AVG(CAST(length_of_stay AS FLOAT)),1) AS avg_los_days,
                       ROUND(AVG(CAST(total_cost_kes AS FLOAT)),0) AS avg_cost_kes,
                       SUM(CASE WHEN discharge_outcome='Deceased' THEN 1 ELSE 0 END) AS deaths
                FROM stg_admissions WHERE ward IS NOT NULL
                GROUP BY ward ORDER BY total_admissions DESC
            """,
            "top_diagnoses": """
                SELECT diagnosis, COUNT(*) AS cases,
                       ROUND(AVG(CAST(length_of_stay AS FLOAT)),1) AS avg_los,
                       ROUND(SUM(CAST(total_cost_kes AS FLOAT)),0) AS total_cost
                FROM stg_admissions WHERE diagnosis IS NOT NULL
                GROUP BY diagnosis ORDER BY cases DESC LIMIT 10
            """,
            "monthly_admissions": """
                SELECT strftime('%Y-%m', admission_date) AS month,
                       COUNT(*) AS admissions,
                       ROUND(AVG(CAST(total_cost_kes AS FLOAT)),0) AS avg_cost
                FROM stg_admissions WHERE admission_date IS NOT NULL
                GROUP BY month ORDER BY month
            """,
            "lab_abnormal_rate": """
                SELECT test_name,
                       COUNT(*) AS total_tests,
                       SUM(CASE WHEN is_abnormal=1 THEN 1 ELSE 0 END) AS abnormal,
                       ROUND(100.0*SUM(CASE WHEN is_abnormal=1 THEN 1 ELSE 0 END)/COUNT(*),1) AS abnormal_pct,
                       ROUND(AVG(CAST(turnaround_hours AS FLOAT)),1) AS avg_tat_hours
                FROM stg_lab_tests WHERE test_name IS NOT NULL
                GROUP BY test_name ORDER BY abnormal_pct DESC
            """,
            "opd_by_clinic": """
                SELECT clinic, COUNT(*) AS visits,
                       ROUND(AVG(CAST(wait_time_mins AS FLOAT)),0) AS avg_wait_mins,
                       ROUND(AVG(CAST(consultation_fee_kes AS FLOAT)),0) AS avg_fee
                FROM stg_outpatient_visits WHERE clinic IS NOT NULL
                GROUP BY clinic ORDER BY visits DESC
            """,
            "patient_demographics": """
                SELECT gender, county,
                       COUNT(*) AS patients,
                       ROUND(AVG(CAST(age_years AS FLOAT)),1) AS avg_age
                FROM stg_patients WHERE gender IS NOT NULL
                GROUP BY gender, county ORDER BY patients DESC
            """,
            "doctor_workload": """
                SELECT attending_doctor,
                       COUNT(*) AS total_encounters,
                       ROUND(AVG(CAST(total_cost_kes AS FLOAT)),0) AS avg_cost
                FROM stg_admissions WHERE attending_doctor IS NOT NULL
                GROUP BY attending_doctor ORDER BY total_encounters DESC
            """,
            "readmission_risk": """
                WITH patient_adm AS (
                    SELECT patient_id, COUNT(*) AS admission_count,
                           ROUND(AVG(CAST(length_of_stay AS FLOAT)),1) AS avg_los,
                           ROUND(SUM(CAST(total_cost_kes AS FLOAT)),0) AS total_spend
                    FROM stg_admissions GROUP BY patient_id
                )
                SELECT p.county, COUNT(DISTINCT pa.patient_id) AS patients,
                       ROUND(AVG(pa.admission_count),2) AS avg_admissions,
                       SUM(CASE WHEN pa.admission_count > 1 THEN 1 ELSE 0 END) AS readmitted
                FROM patient_adm pa
                JOIN stg_patients p ON pa.patient_id = p.patient_id
                WHERE p.county IS NOT NULL
                GROUP BY p.county ORDER BY avg_admissions DESC
            """,
            "revenue_summary": """
                SELECT 'Inpatient' AS source,
                       ROUND(SUM(CAST(total_cost_kes AS FLOAT)),0) AS total_revenue,
                       COUNT(*) AS encounters
                FROM stg_admissions WHERE total_cost_kes IS NOT NULL
                UNION ALL
                SELECT 'Outpatient',
                       ROUND(SUM(CAST(consultation_fee_kes AS FLOAT)),0),
                       COUNT(*)
                FROM stg_outpatient_visits WHERE consultation_fee_kes IS NOT NULL
            """,
            "lab_tat_by_test": """
                SELECT test_name,
                       COUNT(*) AS tests_with_result,
                       ROUND(MIN(CAST(turnaround_hours AS FLOAT)),1) AS min_tat,
                       ROUND(AVG(CAST(turnaround_hours AS FLOAT)),1) AS avg_tat,
                       ROUND(MAX(CAST(turnaround_hours AS FLOAT)),1) AS max_tat
                FROM stg_lab_tests
                WHERE turnaround_hours IS NOT NULL AND turnaround_hours > 0
                GROUP BY test_name ORDER BY avg_tat DESC
            """,
        }

        results = {}
        for name, sql in queries.items():
            try:
                results[name] = pd.read_sql_query(sql, conn)
                self.log.info(f"  Query '{name}': {len(results[name])} rows")
            except Exception as e:
                self.log.warning(f"  Query '{name}' failed: {e}")
                results[name] = pd.DataFrame()
        return results

# ── 4. SQLExporter ────────────────────────────────────────────────────────────

class SQLExporter:
    """Writes PostgreSQL-compatible INSERT statements and analytical views."""

    def __init__(self, log): self.log = log.getChild("SQLExporter")

    def write_inserts(self, dfs, out_dir="sql/cleaning"):
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        for name, df in dfs.items():
            path = f"{out_dir}/load_{name}.sql"
            lines = [f"-- Auto-generated INSERT statements for stg_{name}\n",
                     f"-- Generated: {datetime.now()}\n\n",
                     f"TRUNCATE TABLE health_analytics.stg_{name} CASCADE;\n\n"]
            df2 = df.copy()
            for col in df2.select_dtypes(include=["datetime64[ns]","datetime64[ns, UTC]"]):
                df2[col] = df2[col].dt.strftime("%Y-%m-%d").where(df2[col].notna(), None)
            df2 = df2.where(pd.notna(df2), None)
            for _, row in df2.iterrows():
                vals = []
                for v in row:
                    if v is None or (isinstance(v, float) and np.isnan(v)):
                        vals.append("NULL")
                    elif isinstance(v, (int, float)):
                        vals.append(str(v))
                    elif isinstance(v, bool):
                        vals.append("TRUE" if v else "FALSE")
                    else:
                        safe = str(v).replace("'","''")
                        vals.append(f"'{safe}'")
                cols = ", ".join(df2.columns)
                lines.append(f"INSERT INTO health_analytics.stg_{name} ({cols}) VALUES ({', '.join(vals)});\n")
            with open(path, "w") as f:
                f.writelines(lines[:5])  # write just header sample for portfolio
                f.write(f"-- ... {len(df2)} total INSERT rows generated\n")
            self.log.info(f"  Inserts written → {path}")

    def write_views(self):
        views_sql = '''-- ============================================================
-- ANALYTICAL VIEWS — health_analytics schema
-- Author: Kisavi Shadrack
-- ============================================================
SET search_path TO health_analytics;

-- View 1: Patient summary with admission & visit counts
CREATE OR REPLACE VIEW v_patient_summary AS
SELECT
    p.patient_id,
    p.first_name || \' \' || p.last_name  AS full_name,
    p.age_years,
    p.gender,
    p.county,
    p.blood_group,
    COUNT(DISTINCT a.admission_id)       AS total_admissions,
    COUNT(DISTINCT o.visit_id)           AS total_opd_visits,
    COUNT(DISTINCT l.lab_id)             AS total_lab_tests,
    ROUND(SUM(a.total_cost_kes)::NUMERIC, 0) AS total_inpatient_spend,
    MAX(a.admission_date)                AS last_admission
FROM stg_patients p
LEFT JOIN stg_admissions a       ON p.patient_id = a.patient_id
LEFT JOIN stg_outpatient_visits o ON p.patient_id = o.patient_id
LEFT JOIN stg_lab_tests l        ON p.patient_id = l.patient_id
GROUP BY p.patient_id, p.first_name, p.last_name,
         p.age_years, p.gender, p.county, p.blood_group;

-- View 2: Ward performance KPIs
CREATE OR REPLACE VIEW v_ward_kpis AS
SELECT
    ward,
    COUNT(*)                                          AS total_admissions,
    ROUND(AVG(length_of_stay), 1)                    AS avg_length_of_stay,
    ROUND(AVG(total_cost_kes), 0)                    AS avg_cost_kes,
    SUM(CASE WHEN discharge_outcome = \'Deceased\' THEN 1 ELSE 0 END) AS mortality_count,
    ROUND(100.0 * SUM(CASE WHEN discharge_outcome = \'Deceased\' THEN 1 ELSE 0 END)
          / NULLIF(COUNT(*), 0), 2)                  AS mortality_rate_pct,
    ROUND(SUM(total_cost_kes), 0)                    AS total_revenue
FROM stg_admissions
WHERE ward IS NOT NULL
GROUP BY ward;

-- View 3: Lab test performance
CREATE OR REPLACE VIEW v_lab_performance AS
SELECT
    test_name,
    test_code,
    COUNT(*)                                              AS total_ordered,
    SUM(CASE WHEN result_date IS NOT NULL THEN 1 ELSE 0 END) AS results_returned,
    ROUND(100.0 * SUM(CASE WHEN result_date IS NOT NULL THEN 1 ELSE 0 END)
          / NULLIF(COUNT(*), 0), 1)                       AS completion_rate_pct,
    ROUND(AVG(turnaround_hours), 1)                       AS avg_tat_hours,
    ROUND(100.0 * SUM(CASE WHEN is_abnormal THEN 1 ELSE 0 END)
          / NULLIF(COUNT(*), 0), 1)                       AS abnormal_rate_pct
FROM stg_lab_tests
GROUP BY test_name, test_code;

-- View 4: Monthly activity trend
CREATE OR REPLACE VIEW v_monthly_activity AS
SELECT
    DATE_TRUNC(\'month\', activity_date)  AS month,
    activity_type,
    COUNT(*)                              AS volume,
    ROUND(SUM(revenue), 0)               AS revenue
FROM (
    SELECT admission_date AS activity_date, \'Admission\' AS activity_type, total_cost_kes AS revenue
    FROM stg_admissions WHERE admission_date IS NOT NULL
    UNION ALL
    SELECT visit_date, \'OPD Visit\', consultation_fee_kes
    FROM stg_outpatient_visits WHERE visit_date IS NOT NULL
    UNION ALL
    SELECT ordered_date, \'Lab Test\', NULL
    FROM stg_lab_tests WHERE ordered_date IS NOT NULL
) combined
GROUP BY DATE_TRUNC(\'month\', activity_date), activity_type
ORDER BY month, activity_type;

-- View 5: High-risk patients (multiple admissions + abnormal labs)
CREATE OR REPLACE VIEW v_high_risk_patients AS
WITH adm_counts AS (
    SELECT patient_id, COUNT(*) AS admissions,
           ROUND(AVG(length_of_stay), 1) AS avg_los
    FROM stg_admissions GROUP BY patient_id
),
lab_abnormal AS (
    SELECT patient_id,
           SUM(CASE WHEN is_abnormal THEN 1 ELSE 0 END) AS abnormal_results,
           COUNT(*) AS total_labs
    FROM stg_lab_tests GROUP BY patient_id
)
SELECT p.patient_id, p.first_name || \' \' || p.last_name AS full_name,
       p.age_years, p.gender, p.county,
       COALESCE(a.admissions, 0)        AS admissions,
       COALESCE(a.avg_los, 0)           AS avg_los,
       COALESCE(l.abnormal_results, 0)  AS abnormal_lab_results,
       COALESCE(l.total_labs, 0)        AS total_labs
FROM stg_patients p
LEFT JOIN adm_counts a  ON p.patient_id = a.patient_id
LEFT JOIN lab_abnormal l ON p.patient_id = l.patient_id
WHERE COALESCE(a.admissions, 0) > 1
   OR COALESCE(l.abnormal_results, 0) > 2
ORDER BY admissions DESC, abnormal_lab_results DESC;
'''
        path = "sql/views/analytical_views.sql"
        Path("sql/views").mkdir(parents=True, exist_ok=True)
        with open(path, "w") as f:
            f.write(views_sql)
        self.log.info(f"Views written → {path}")

# ── 5. ChartGenerator ─────────────────────────────────────────────────────────

class ChartGenerator:
    PALETTE = ["#1B4F72","#2E86C1","#28B463","#E74C3C","#F39C12","#7D3C98","#1ABC9C","#E67E22"]

    def __init__(self, log):
        self.log = log.getChild("Charts")
        Path("dashboard/charts").mkdir(parents=True, exist_ok=True)
        plt.rcParams.update({"font.family":"DejaVu Sans","axes.spines.top":False,
                             "axes.spines.right":False,"figure.dpi":150})

    def _save(self, fig, name):
        fig.savefig(f"dashboard/charts/{name}", bbox_inches="tight")
        plt.close(fig)
        self.log.info(f"  Chart saved: {name}")

    def generate_all(self, results):
        self.log.info("Generating charts …")
        self._ward_admissions(results.get("admissions_by_ward", pd.DataFrame()))
        self._top_diagnoses(results.get("top_diagnoses", pd.DataFrame()))
        self._monthly_trend(results.get("monthly_admissions", pd.DataFrame()))
        self._lab_abnormal(results.get("lab_abnormal_rate", pd.DataFrame()))
        self._opd_clinics(results.get("opd_by_clinic", pd.DataFrame()))
        self._revenue_split(results.get("revenue_summary", pd.DataFrame()))

    def _ward_admissions(self, df):
        if df.empty: return
        fig, axes = plt.subplots(1, 2, figsize=(12, 4))
        df = df.sort_values("total_admissions", ascending=True)
        axes[0].barh(df["ward"], df["total_admissions"], color=self.PALETTE[0])
        axes[0].set_title("Admissions by Ward", fontweight="bold")
        axes[0].set_xlabel("Admissions")
        axes[1].bar(df["ward"], df["avg_los_days"], color=self.PALETTE[1], edgecolor="white")
        axes[1].set_title("Avg Length of Stay (Days)", fontweight="bold")
        axes[1].set_ylabel("Days")
        plt.xticks(rotation=30, ha="right")
        plt.tight_layout()
        self._save(fig, "01_ward_admissions.png")

    def _top_diagnoses(self, df):
        if df.empty: return
        fig, ax = plt.subplots(figsize=(10, 5))
        df = df.sort_values("cases")
        bars = ax.barh(df["diagnosis"], df["cases"], color=self.PALETTE[2], edgecolor="white")
        ax.set_title("Top 10 Diagnoses by Admission Count", fontweight="bold", fontsize=13)
        ax.set_xlabel("Cases")
        for bar in bars:
            ax.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                    f"{int(bar.get_width())}", va="center", fontsize=9)
        plt.tight_layout()
        self._save(fig, "02_top_diagnoses.png")

    def _monthly_trend(self, df):
        if df.empty: return
        fig, ax1 = plt.subplots(figsize=(12, 4))
        ax2 = ax1.twinx()
        ax1.bar(df["month"], df["admissions"], color=self.PALETTE[0], alpha=0.7, label="Admissions")
        ax2.plot(df["month"], df["avg_cost"]/1000, color=self.PALETTE[3], marker="o",
                 linewidth=2, label="Avg Cost (K)")
        ax1.set_title("Monthly Admissions & Average Cost Trend", fontweight="bold", fontsize=13)
        ax1.set_ylabel("Admissions", color=self.PALETTE[0])
        ax2.set_ylabel("Avg Cost (KES Thousands)", color=self.PALETTE[3])
        plt.xticks(rotation=45, ha="right")
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1+lines2, labels1+labels2, loc="upper left")
        plt.tight_layout()
        self._save(fig, "03_monthly_trend.png")

    def _lab_abnormal(self, df):
        if df.empty: return
        df = df.sort_values("abnormal_pct", ascending=True)
        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.barh(df["test_name"], df["abnormal_pct"],
                       color=[self.PALETTE[3] if x > 40 else self.PALETTE[1] for x in df["abnormal_pct"]])
        ax.set_title("Abnormal Result Rate by Lab Test (%)", fontweight="bold", fontsize=13)
        ax.set_xlabel("Abnormal Rate (%)")
        ax.axvline(40, color="red", linestyle="--", linewidth=1, alpha=0.5, label="40% threshold")
        ax.legend()
        for bar in bars:
            ax.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                    f"{bar.get_width():.1f}%", va="center", fontsize=9)
        plt.tight_layout()
        self._save(fig, "04_lab_abnormal_rates.png")

    def _opd_clinics(self, df):
        if df.empty: return
        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        df = df.sort_values("visits", ascending=False)
        axes[0].bar(df["clinic"], df["visits"], color=self.PALETTE[4], edgecolor="white")
        axes[0].set_title("OPD Visits by Clinic", fontweight="bold")
        axes[0].set_ylabel("Visits")
        axes[0].tick_params(axis="x", rotation=35)
        axes[1].bar(df["clinic"], df["avg_wait_mins"], color=self.PALETTE[5], edgecolor="white")
        axes[1].set_title("Average Wait Time by Clinic (mins)", fontweight="bold")
        axes[1].set_ylabel("Minutes")
        axes[1].tick_params(axis="x", rotation=35)
        plt.tight_layout()
        self._save(fig, "05_opd_clinics.png")

    def _revenue_split(self, df):
        if df.empty: return
        fig, ax = plt.subplots(figsize=(6, 6))
        wedges, texts, autotexts = ax.pie(
            df["total_revenue"], labels=df["source"],
            autopct="%1.1f%%", colors=[self.PALETTE[0], self.PALETTE[2]],
            startangle=90, pctdistance=0.8, wedgeprops=dict(edgecolor="white", linewidth=2))
        for at in autotexts: at.set_fontsize(12)
        ax.set_title("Revenue: Inpatient vs Outpatient", fontweight="bold", fontsize=13)
        self._save(fig, "06_revenue_split.png")

# ── 6. ReportGenerator ────────────────────────────────────────────────────────

class ReportGenerator:
    BLUE = "1B4F72"; LBLUE = "D6EAF8"; GREEN = "1E8449"; RED = "C0392B"

    def __init__(self, log): self.log = log.getChild("Reports")

    def _hdr(self, cell, bg, fg="FFFFFF"):
        cell.font = Font(bold=True, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _border(self, ws, r1, c1, r2, c2):
        t = Side(style="thin")
        for row in ws.iter_rows(r1, r2, c1, c2):
            for c in row:
                c.border = Border(left=t, right=t, top=t, bottom=t)

    def _autowidth(self, ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 45)

    def write_summary(self, results, stats):
        Path("reports").mkdir(exist_ok=True)
        wb = openpyxl.Workbook()

        # Sheet 1 — Pipeline Stats
        ws = wb.active; ws.title = "Pipeline Stats"
        ws.merge_cells("A1:D1")
        ws["A1"] = "Health Analytics Pipeline — Run Summary"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=self.BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26
        ws.append([])
        ws.append(["Table", "Raw Rows", "Clean Rows", "Removed"])
        for j in range(1,5): self._hdr(ws.cell(3,j), self.BLUE)
        for name, s in stats.tables.items():
            ws.append([name, s["raw"], s["clean"], s["raw"]-s["clean"]])
        self._border(ws,3,1,ws.max_row,4); self._autowidth(ws)

        # Sheet 2 — Ward KPIs
        self._df_sheet(wb, results.get("admissions_by_ward", pd.DataFrame()),
                       "Ward KPIs", "Ward Performance Summary")
        # Sheet 3 — Top Diagnoses
        self._df_sheet(wb, results.get("top_diagnoses", pd.DataFrame()),
                       "Top Diagnoses", "Top 10 Diagnoses by Admissions")
        # Sheet 4 — Lab Performance
        self._df_sheet(wb, results.get("lab_abnormal_rate", pd.DataFrame()),
                       "Lab Performance", "Lab Test Abnormal Rates & Turnaround")
        # Sheet 5 — OPD Clinics
        self._df_sheet(wb, results.get("opd_by_clinic", pd.DataFrame()),
                       "OPD Clinics", "Outpatient Clinic Performance")
        # Sheet 6 — Revenue
        self._df_sheet(wb, results.get("revenue_summary", pd.DataFrame()),
                       "Revenue", "Revenue Summary by Source")

        wb.save("reports/summary_report.xlsx")
        self.log.info("Summary report → reports/summary_report.xlsx")

    def _df_sheet(self, wb, df, sheet_name, title):
        ws = wb.create_sheet(sheet_name)
        n = len(df.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n,1))
        ws.cell(1,1).value = title
        ws.cell(1,1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(1,1).fill = PatternFill("solid", fgColor=self.BLUE)
        ws.cell(1,1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24
        ws.append([])
        if df.empty:
            ws.append(["No data"]); return
        ws.append(list(df.columns))
        for j in range(1, n+1): self._hdr(ws.cell(3,j), self.BLUE)
        for i, (_, row) in enumerate(df.iterrows()):
            ws.append([round(v,2) if isinstance(v,float) else v for v in row])
            fill = self.LBLUE if i%2==0 else "FFFFFF"
            for j in range(1,n+1):
                ws.cell(ws.max_row,j).fill = PatternFill("solid",fgColor=fill)
        self._border(ws,3,1,ws.max_row,n); self._autowidth(ws)

# ── Pipeline Orchestrator ─────────────────────────────────────────────────────

class Pipeline:
    def __init__(self):
        self.log = setup_logging("logs/pipeline.log")
        self.stats = PipelineStats()

    def run(self):
        log = self.log
        log.info("="*60)
        log.info("  Hospital Health Analytics Pipeline  v1.0")
        log.info("  Author: Kisavi Shadrack")
        log.info("="*60)

        # Load
        loader = DataLoader(log)
        raw = loader.load_all()

        # Clean
        cleaner = DataCleaner(log, self.stats)
        p_clean  = cleaner.clean_patients(raw["patients"])
        valid_pids = set(p_clean["patient_id"].dropna())
        a_clean  = cleaner.clean_admissions(raw["admissions"], valid_pids)
        l_clean  = cleaner.clean_lab_tests(raw["lab_tests"], valid_pids)
        o_clean  = cleaner.clean_outpatient(raw["outpatient_visits"], valid_pids)

        clean_dfs = {"stg_patients": p_clean, "stg_admissions": a_clean,
                     "stg_lab_tests": l_clean, "stg_outpatient_visits": o_clean}

        # Save clean CSVs
        Path("output").mkdir(exist_ok=True)
        for name, df in clean_dfs.items():
            df.to_csv(f"output/{name}.csv", index=False)
        log.info("Clean CSVs saved to output/")

        # SQL Export
        exporter = SQLExporter(log)
        exporter.write_inserts(clean_dfs)
        exporter.write_views()

        # Analytics
        engine = AnalyticsEngine(log)
        conn = engine.load_to_sqlite(clean_dfs)
        results = engine.run(conn)

        # Charts
        charter = ChartGenerator(log)
        charter.generate_all(results)

        # Reports
        reporter = ReportGenerator(log)
        reporter.write_summary(results, self.stats)

        log.info("="*60)
        log.info("Pipeline complete!")
        for t, s in self.stats.tables.items():
            log.info(f"  {t:25s}: {s['raw']:>4} → {s['clean']:>4} rows")
        log.info("="*60)

        return results, clean_dfs

if __name__ == "__main__":
    Pipeline().run()
